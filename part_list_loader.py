"""
Part List 폴더의 BOM 엑셀에서 LOCATION → 정규화 품목명 인덱스 생성.
- IC / MOSFET / TR / DIODE / EL-CAP(C-EL) 위주
- 반도체: 세미콜론 뒤 MPN 우선 (예: TR-SMALL SIGNAL;2SC4541U → 2SC4541U)
- C-EL: C EL 접두·]·온도 표기 등 정리 (예: NXH 35V 330uF 10*12.5 형태로 근접)
"""
from __future__ import annotations

import os
import re
import pickle
import hashlib
from typing import Callable, Dict, List, Optional, Tuple

from openpyxl import load_workbook

# 헤더 셀에서 찾을 키워드 (대소문자 무시)
HDR_LOCATION = "LOCATION"
HDR_PART_NO = "PART NO"
HDR_DESC = "DESCRIPTION"
HDR_SPEC = "SPECIFICATION"
HDR_VENDOR = "VENDOR"


def _norm_ws(s: str) -> str:
    return " ".join(str(s).split()).strip()


def _pack_candidate(vendor: str, norm_text: str) -> str:
    v = _norm_ws(vendor or "")
    n = _norm_ws(norm_text or "")
    if v and n:
        return f"{v}||{n}"
    return n


def _unpack_candidate(cand: str) -> Tuple[str, str]:
    s = str(cand or "")
    if "||" in s:
        v, n = s.split("||", 1)
        return _norm_ws(v), _norm_ws(n)
    return "", _norm_ws(s)


def classify_pl_line(blob: str) -> Optional[str]:
    """IC/MOSFET/TR/DIODE/EL_CAP 중 하나 또는 None."""
    u = blob.upper()
    if re.search(r"\bC\s*[-]?\s*EL\b", blob, re.I) or "EL-CAP" in u or "E-CAP" in u:
        return "EL_CAP"
    # C-EL 명시가 없어도 전해콘덴서 패턴(전압+용량)으로 분류
    if re.search(r"\d+(?:\.\d+)?\s*V\b", u) and re.search(r"\d+(?:\.\d+)?\s*(UF|ΜF|μF)\b", u):
        return "EL_CAP"
    if "MOSFET" in u or ("FET" in u and "C-FILM" not in u and "C FILM" not in u):
        return "MOSFET"
    if "RECTIFIER" in u or "DIODE" in u or "D-RECT" in u:
        return "DIODE"
    if "IC-PFC" in u or "IC-PWM" in u or re.search(r"\bIC[- ]", u):
        return "IC"
    if "SMALL SIGNAL" in u or "TR-SMALL" in u or re.search(r"\bTR[- ]", u):
        return "TR"
    return None


def normalize_semiconductor_mpn(part_no: str, desc: str, spec: str) -> str:
    """세미콜론 뒤 품번 우선."""
    for raw in (part_no, desc, spec):
        if not raw:
            continue
        s = str(raw).strip()
        if ";" in s:
            return _norm_ws(s.split(";")[-1])
    return _norm_ws(str(part_no or ""))


def normalize_el_cap_line(part_no: str, desc: str, spec: str) -> str:
    """C-EL 계열: PL 전체에서 쓰는 짧은 스펙 문자열."""
    blob = " ".join(filter(None, [str(part_no or ""), str(desc or ""), str(spec or "")]))
    t = blob.replace("]", " ")
    t = _norm_ws(t)

    # 벤더명 추출: ';' 뒤 토큰 우선, 없으면 일반 토큰 탐색
    stop_vendor = {
        "EL", "C", "CAP", "CE", "CEL", "E", "ECAP", "ELCAP",
        "V", "UF", "PF", "NF", "MM", "DIA", "RIPPLE", "LOW", "ESR",
    }
    vendor = ""
    for raw in (str(part_no or ""), str(desc or ""), str(spec or "")):
        s = raw.strip()
        if not s:
            continue
        if ";" in s:
            cand = s.split(";")[-1].strip()
            tok = re.split(r"[\s,/]+", cand)[0].strip().upper() if cand else ""
            if tok and tok not in stop_vendor and re.search(r"[A-Z]", tok):
                vendor = tok
                break
    if not vendor:
        for tok in re.findall(r"\b[A-Z][A-Z0-9]{1,}\b", t.upper()):
            if tok in stop_vendor:
                continue
            # 전압/단위 토큰 배제
            if re.fullmatch(r"\d+(V|UF|PF|NF|MM)?", tok):
                continue
            vendor = tok
            break

    v_match = re.search(r"(\d+(?:\.\d+)?)\s*V\b", t, re.I)
    cap_match = re.search(r"(\d+(?:\.\d+)?)\s*(uF|UF|μF|ΜF)\b", t, re.I)
    size_match = re.search(r"(\d+(?:\.\d+)?)\s*[*xX]\s*(\d+(?:\.\d+)?)", t)

    parts = []
    if vendor:
        parts.append(vendor)
    if v_match:
        parts.append(f"{v_match.group(1)}V")
    if cap_match:
        parts.append(f"{cap_match.group(1)}uF")
    if size_match:
        parts.append(f"{size_match.group(1)}*{size_match.group(2)}")

    if parts:
        return _norm_ws(" ".join(parts))
    return t


def normalize_pl_row(part_no: str, desc: str, spec: str) -> str:
    blob = " ".join(filter(None, [str(part_no or ""), str(desc or ""), str(spec or "")]))
    cat = classify_pl_line(blob)
    if cat == "EL_CAP":
        return normalize_el_cap_line(part_no, desc, spec)
    return normalize_semiconductor_mpn(part_no, desc, spec)


def _find_header_row(ws, max_row_scan: int = 30, max_col_scan: int = 30) -> Optional[Tuple[int, Dict[str, int]]]:
    """(header_row_1based, {LOCATION: col_idx, ...}) 또는 None."""
    mr = min(ws.max_row or 1, max_row_scan)
    mc = min(ws.max_column or 1, max_col_scan)
    for r in range(1, mr + 1):
        labels = {}
        for c in range(1, mc + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            key = str(v).strip().upper().replace(".", "")
            if HDR_LOCATION in key and "LOCATION" in key:
                labels["LOCATION"] = c
            if "PART" in key and "NO" in key:
                labels["PART_NO"] = c
            if key == "DESCRIPTION" or key.startswith("DESCRIPTION"):
                labels["DESCRIPTION"] = c
            if "SPECIFICATION" in key or key == "SPEC":
                labels["SPECIFICATION"] = c
            if "VENDOR" in key or "MAKER" in key or "SUPPLIER" in key:
                labels["VENDOR"] = c
        if "LOCATION" in labels and "PART_NO" in labels:
            return r, labels
    return None


def _should_skip_sheet(name: str) -> bool:
    """리비전 비교 전용 시트만 제외 (단일 Sheet1 BOM은 유지)."""
    ln = name.lower()
    if "rev history" in ln or "revision history" in ln:
        return True
    return False


def parse_workbook(
    path: str,
    log: Optional[Callable[[str], None]] = None,
    progress_cb: Optional[Callable[[int, int, str], None]] = None,
) -> Dict[str, List[str]]:
    """
    단일 Part List xlsx → {로케이션대문자: [정규화품명, ...]}
    동일 Loc 여러 행(벤더 대체)은 리스트에 순서대로 누적.
    """
    out: Dict[str, List[str]] = {}
    out_seen: Dict[str, set] = {}
    wb = load_workbook(path, read_only=True, data_only=True)
    base = os.path.basename(path)
    try:
        sheets = [s for s in wb.sheetnames if not _should_skip_sheet(s)]
        # 시트별 동일 가중치(대략적인 진행률)
        total_units = max(1, len(sheets) * 1000)
        done_units = 0
        if progress_cb:
            progress_cb(done_units, total_units, f"{base} 시작")
        for sheet_name in sheets:
            ws = wb[sheet_name]
            mc = ws.max_column or 0
            if mc > 400:
                if log:
                    log(f"[PL] {base} :: 시트 '{sheet_name}' 열 수 과다({mc}) — 건너뜀 (병합/서식으로 열이 비정상 확장된 경우)")
                done_units += 1000
                if progress_cb:
                    progress_cb(done_units, total_units, f"{base}::{sheet_name} skip")
                continue
            found = _find_header_row(ws, max_row_scan=25, max_col_scan=min(24, mc or 24))
            if not found:
                done_units += 1000
                if progress_cb:
                    progress_cb(done_units, total_units, f"{base}::{sheet_name} no-header")
                continue
            hdr_r, cols = found
            loc_c = cols["LOCATION"]
            pno_c = cols["PART_NO"]
            desc_c = cols.get("DESCRIPTION")
            spec_c = cols.get("SPECIFICATION")
            vendor_c = cols.get("VENDOR")
            max_r = min(ws.max_row or hdr_r, 2000)
            max_c_needed = max([c for c in [loc_c, pno_c, desc_c or 0, spec_c or 0, vendor_c or 0] if c])
            sheet_hits = 0
            last_loc_str = ""
            row_total = max(1, max_r - hdr_r)
            # openpyxl read_only에서 ws.cell 반복 호출이 느릴 수 있어, 필요한 col만 values_only로 읽는다.
            row_iter = ws.iter_rows(
                min_row=hdr_r + 1,
                max_row=max_r,
                min_col=1,
                max_col=max_c_needed,
                values_only=True,
            )
            for idx0, row_vals in enumerate(row_iter, start=1):
                # idx0는 header 아래부터 1
                loc_val = row_vals[loc_c - 1] if loc_c <= len(row_vals) else None
                pno = row_vals[pno_c - 1] if pno_c <= len(row_vals) else None
                if loc_val is None and (pno is None or str(pno).strip() == ""):
                    if progress_cb and idx0 % 25 == 0:
                        done = done_units + int((idx0 / row_total) * 1000)
                        progress_cb(done, total_units, f"{base}::{sheet_name}")
                    continue
                desc = row_vals[desc_c - 1] if desc_c and desc_c <= len(row_vals) else ""
                spec = row_vals[spec_c - 1] if spec_c and spec_c <= len(row_vals) else ""
                vendor = row_vals[vendor_c - 1] if vendor_c and vendor_c <= len(row_vals) else ""
                blob = " ".join(
                    filter(None, [str(loc_val or ""), str(pno or ""), str(desc or ""), str(spec or ""), str(vendor or "")])
                )
                cat = classify_pl_line(blob)
                if cat not in ("IC", "MOSFET", "TR", "DIODE", "EL_CAP"):
                    continue
                norm = normalize_pl_row(
                    str(pno or ""),
                    str(desc or ""),
                    str(spec or ""),
                )
                if not norm:
                    continue
                packed = _pack_candidate(str(vendor or ""), norm)
                if not packed:
                    continue
                loc_str = str(loc_val or "").strip()
                if loc_str:
                    last_loc_str = loc_str
                elif last_loc_str:
                    # BOM에서 대체 벤더 행은 LOCATION이 병합되어 비어있을 수 있음
                    loc_str = last_loc_str
                if not loc_str or loc_str == "-":
                    continue
                sheet_hits += 1
                for piece in re.split(r"[,;]\s*|\s+/\s+", loc_str):
                    piece = piece.strip().upper()
                    if len(piece) < 3:
                        continue
                    out.setdefault(piece, [])
                    if piece not in out_seen:
                        out_seen[piece] = set()
                    if packed not in out_seen[piece]:
                        out_seen[piece].add(packed)
                        out[piece].append(packed)
                if progress_cb and idx0 % 25 == 0:
                    done = done_units + int((idx0 / row_total) * 1000)
                    progress_cb(done, total_units, f"{base}::{sheet_name}")
            done_units += 1000
            if progress_cb:
                progress_cb(done_units, total_units, f"{base}::{sheet_name} done")
            if sheet_hits and log:
                log(f"[PL] {base} :: '{sheet_name}' 대상 {sheet_hits}행 (누적 Loc {len(out)})")
    finally:
        wb.close()
    return out


def merge_pl_dicts(
    a: Dict[str, List[str]], b: Dict[str, List[str]]
) -> Dict[str, List[str]]:
    for k, lst in b.items():
        a.setdefault(k, [])
        for x in lst:
            if x not in a[k]:
                a[k].append(x)
    return a


def _merge_pl_dicts_fast(
    a: Dict[str, List[str]],
    a_seen: Dict[str, set],
    b: Dict[str, List[str]],
) -> None:
    """a에 b를 합치되, membership은 set으로 가속."""
    for k, lst in b.items():
        if k not in a:
            a[k] = []
            a_seen[k] = set()
        seen = a_seen.setdefault(k, set())
        for x in lst:
            if x in seen:
                continue
            seen.add(x)
            a[k].append(x)


def _file_fingerprint(path: str) -> tuple[str, int, int]:
    st = os.stat(path)
    return (os.path.basename(path), int(st.st_size), int(st.st_mtime))


def _compute_cache_key(kind: str, paths: List[str]) -> str:
    h = hashlib.sha256()
    h.update(kind.encode("utf-8", "ignore"))
    for p in sorted(paths, key=lambda s: s.lower()):
        try:
            bn, sz, mt = _file_fingerprint(p)
        except OSError:
            bn, sz, mt = (os.path.basename(p), 0, 0)
        h.update(b"\0")
        h.update(bn.encode("utf-8", "ignore"))
        h.update(str(sz).encode("ascii"))
        h.update(str(mt).encode("ascii"))
    return h.hexdigest()


def _cache_paths(base_dir: str, cache_key: str) -> str:
    cache_dir = os.path.join(base_dir, "__pl_cache__")
    os.makedirs(cache_dir, exist_ok=True)
    return os.path.join(cache_dir, f"pl_index_{cache_key}.pkl")


def load_part_list_index(
    base_dir: str,
    log: Optional[Callable[[str], None]] = None,
    progress_cb: Optional[Callable[[int, int, str], None]] = None,
) -> Dict[str, List[str]]:
    """
    base_dir/Part List 아래 모든 .xlsx 스캔 후 Loc → 정규화 품명 리스트.
    """
    folder = os.path.join(base_dir, "Part List")
    if not os.path.isdir(folder):
        if log:
            log(f"[PL] 폴더 없음: {folder}")
        return {}
    merged: Dict[str, List[str]] = {}
    files = sorted(
        f for f in os.listdir(folder) if f.lower().endswith(".xlsx") and not f.startswith("~$")
    )
    full_paths = [os.path.join(folder, fn) for fn in files]
    cache_key = _compute_cache_key("folder", full_paths)
    cache_file = _cache_paths(base_dir, cache_key)
    if os.path.isfile(cache_file):
        try:
            with open(cache_file, "rb") as f:
                cached = pickle.load(f)
            if isinstance(cached, dict):
                if log:
                    log(f"[PL] 캐시 로드: {os.path.basename(cache_file)} (파일 {len(files)}개)")
                return cached
        except Exception:
            pass
    total = len(files)
    if progress_cb:
        progress_cb(0, total, "시작")
    merged: Dict[str, List[str]] = {}
    merged_seen: Dict[str, set] = {}
    for idx, path in enumerate(full_paths, start=1):
        fn = os.path.basename(path)
        try:
            def _inner(done, total_inner, name):
                # 파일 단위 진행률을 파일 내부 진행률로 세분화
                global_done = (idx - 1) + ((done / total_inner) if total_inner else 1.0)
                progress_cb(global_done, total, name) if progress_cb else None
            part = parse_workbook(path, log=log, progress_cb=_inner if progress_cb else None)
            _merge_pl_dicts_fast(merged, merged_seen, part)
        except Exception as e:
            if log:
                log(f"[PL] 건너뜀 {fn}: {e}")
        if progress_cb:
            progress_cb(idx, total, fn)
    if log:
        nloc = len(merged)
        nent = sum(len(v) for v in merged.values())
        log(f"[PL] 로드 완료: 파일 {len(files)}개 → Loc {nloc}개, 후보 {nent}건")
    try:
        with open(cache_file, "wb") as f:
            pickle.dump(merged, f, protocol=pickle.HIGHEST_PROTOCOL)
        if log:
            log(f"[PL] 캐시 저장: {os.path.basename(cache_file)}")
    except Exception:
        pass
    return merged


def load_part_list_from_paths(
    paths: List[str],
    log: Optional[Callable[[str], None]] = None,
    progress_cb: Optional[Callable[[int, int, str], None]] = None,
) -> Dict[str, List[str]]:
    """
    사용자가 선택한 PL 파일 목록에서 Loc → 정규화 품명 리스트 생성.
    """
    valid = [
        p for p in paths
        if p and os.path.isfile(p) and p.lower().endswith((".xlsx", ".xlsm")) and not os.path.basename(p).startswith("~$")
    ]
    # cache는 첫 path의 base_dir(프로젝트 루트) 기준으로 둔다.
    base_dir = os.path.dirname(os.path.abspath(valid[0])) if valid else os.getcwd()
    cache_key = _compute_cache_key("paths", valid)
    cache_file = _cache_paths(base_dir, cache_key)
    if os.path.isfile(cache_file):
        try:
            with open(cache_file, "rb") as f:
                cached = pickle.load(f)
            if isinstance(cached, dict):
                if log:
                    log(f"[PL] 선택 파일 캐시 로드: {os.path.basename(cache_file)} (파일 {len(valid)}개)")
                return cached
        except Exception:
            pass
    merged: Dict[str, List[str]] = {}
    merged_seen: Dict[str, set] = {}
    total = len(valid)
    if progress_cb:
        progress_cb(0, total, "시작")
    for idx, path in enumerate(valid, start=1):
        try:
            def _inner(done, total_inner, name):
                global_done = (idx - 1) + ((done / total_inner) if total_inner else 1.0)
                progress_cb(global_done, total, name) if progress_cb else None
            part = parse_workbook(path, log=log, progress_cb=_inner if progress_cb else None)
            _merge_pl_dicts_fast(merged, merged_seen, part)
        except Exception as e:
            if log:
                log(f"[PL] 건너뜀 {os.path.basename(path)}: {e}")
        if progress_cb:
            progress_cb(idx, total, os.path.basename(path))
    if log:
        nloc = len(merged)
        nent = sum(len(v) for v in merged.values())
        log(f"[PL] 선택 파일 로드 완료: 파일 {len(valid)}개 → Loc {nloc}개, 후보 {nent}건")
    try:
        with open(cache_file, "wb") as f:
            pickle.dump(merged, f, protocol=pickle.HIGHEST_PROTOCOL)
        if log:
            log(f"[PL] 선택 파일 캐시 저장: {os.path.basename(cache_file)}")
    except Exception:
        pass
    return merged


def enrich_m_values_from_pl(
    m_values: List[str],
    pl_index: Dict[str, List[str]],
    vendor_rank_by_loc: Optional[Dict[str, int]],
    log_func: Callable[[str], None],
) -> List[str]:
    """M~R에서 나온 문자열에 PL에서 품명 후보를 덧붙여 DB 매칭률을 높임."""
    if not pl_index:
        return m_values
    extra: List[str] = []
    seen = set()
    for m_str in m_values:
        for tok in re.split(r"[\s,;/]+", m_str):
            tok = tok.strip()
            if len(tok) < 3:
                continue
            key = tok.upper()
            if key not in pl_index:
                continue
            cands = list(pl_index[key])
            pref_rank = 0
            if vendor_rank_by_loc:
                pref_rank = int(vendor_rank_by_loc.get(key, 0) or 0)
            if pref_rank > 0 and len(cands) >= pref_rank:
                # C안: 로케이션별로 지정된 N차 벤더만 우선 사용
                cands = [cands[pref_rank - 1]]
            for cand in cands:
                vendor, norm_text = _unpack_candidate(cand)
                if not norm_text:
                    continue
                low = norm_text.lower()
                if low in seen:
                    continue
                seen.add(low)
                extra.append(norm_text)
                if pref_rank > 0:
                    if vendor:
                        log_func(f"  [PL] Loc '{tok}' ({pref_rank}차, {vendor}) → 검색어 추가 '{norm_text}'")
                    else:
                        log_func(f"  [PL] Loc '{tok}' ({pref_rank}차) → 검색어 추가 '{norm_text}'")
                else:
                    if vendor:
                        log_func(f"  [PL] Loc '{tok}' ({vendor}) → 검색어 추가 '{norm_text}'")
                    else:
                        log_func(f"  [PL] Loc '{tok}' → 검색어 추가 '{norm_text}'")
    return m_values + extra
