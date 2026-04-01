"""
Excel Part Number Matcher
- 엑셀1 파일을 드래그앤드롭으로 불러옴
- Database1.xlsx와 Part Number 비교
- 일치하는 spec 값을 엑셀1의 특정 위치에 기입
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import os
import sys
import shutil
import threading
import queue
import json
from datetime import datetime
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.units import points_to_pixels
from openpyxl.drawing.image import Image as XLImage
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

try:
    from PIL import Image, ImageTk
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

try:
    from part_list_loader import (
        load_part_list_index,
        load_part_list_from_paths,
        enrich_m_values_from_pl,
    )
    HAS_PART_LIST = True
except ImportError:
    HAS_PART_LIST = False
    def load_part_list_index(base_dir, log=None, progress_cb=None):
        return {}
    def load_part_list_from_paths(paths, log=None, progress_cb=None):
        return {}
    def enrich_m_values_from_pl(m_values, pl_index, vendor_rank_by_loc, log_func):
        return m_values

# ─────────────────────────────────────────────
# 경로 설정
# ─────────────────────────────────────────────
# PyInstaller frozen 환경에서는 리소스가 sys._MEIPASS 아래에 풀립니다.
# 일반 실행 시에는 __file__ 기준 경로를 사용합니다.
BASE_DIR = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
DB_DIR   = os.path.join(BASE_DIR, "DB")
DB_FILE  = os.path.join(DB_DIR, "Database.xlsx")
STRESS_RULE_FILE = os.path.join(DB_DIR, "Stress Rule.xlsx")
CONFIG_FILE = os.path.join(BASE_DIR, "app_config.json")
RECENT_MAX = 12
LOG_HISTORY_MAX = 3000

def _default_config():
    return {
        "recent_files": [],
        "open_result_after_run": True,
        "normalize_cell_text": True,
        "validate_workbook_on_run": True,
        "dry_run": False,
        "use_part_list_lookup": True,
    }

def load_app_config():
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            base = _default_config()
            base.update({k: data[k] for k in base if k in data})
            if "recent_files" in data:
                base["recent_files"] = data["recent_files"][:RECENT_MAX]
            return base
    except Exception:
        pass
    return _default_config()

def save_app_config(cfg):
    try:
        cfg = dict(cfg)
        cfg["recent_files"] = [p for p in cfg.get("recent_files", []) if p and os.path.isfile(p)][:RECENT_MAX]
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def add_recent_file(cfg, path):
    if not path or not os.path.isfile(path):
        return cfg
    path = os.path.normpath(path)
    lst = [p for p in cfg.get("recent_files", []) if os.path.normpath(p) != path]
    lst.insert(0, path)
    cfg["recent_files"] = lst[:RECENT_MAX]
    save_app_config(cfg)
    return cfg

def normalize_cell_text(s):
    """M~R 등 셀에서 읽은 문자열 전처리: 공백·개행 정리."""
    if s is None:
        return ""
    t = str(s).replace("\r\n", "\n").replace("\r", "\n")
    t = " ".join(t.split())
    return t.strip()

def normalize_part_key(s):
    """중복 비교용 키 (공백·하이픈 제거, 대문자)."""
    if not s:
        return ""
    return re.sub(r"[\s\-_]+", "", str(s).strip()).upper()


def _strip_cap_ripple_brackets_db(s):
    """CAP 리플 전류 표기에서 '[100Khz]' 등 대괄호 블록 제거."""
    if s is None:
        return ""
    t = str(s).strip()
    t = re.sub(r"\s*\[[^\]]+\]", "", t)
    return " ".join(t.split()).strip()


def _cap_db_d_key(v) -> str:
    t = str(v or "").strip().lower().replace("↓", "").replace("↑", "")
    t = " ".join(t.split())
    return re.sub(r"\s+", "", t)


def _cap_db_e_key(v) -> str:
    return _cap_db_d_key(_strip_cap_ripple_brackets_db(v))


def cap_import_weak_dup_key(part_name: str, specs: list) -> tuple:
    """CAP: 시리즈(첫 토큰) + D열 + E열(리플·주파수 제거) — 불완전 행 재추가 방지."""
    s = str(part_name or "").strip()
    tok = s.split()[0] if s else ""
    sk = normalize_part_key(tok)
    if len(specs) < 2:
        return (sk, "", "")
    return (sk, _cap_db_d_key(specs[0]), _cap_db_e_key(specs[1]))


def cap_part_row_is_incomplete(part_name: str) -> bool:
    """CAP C열에 uF 또는 치수( n*n )가 없으면 불완전."""
    s = str(part_name or "")
    if not re.search(r"\d+(?:\.\d+)?\s*(?:uF|UF|μF|ΜF)", s, re.I):
        return True
    if not re.search(r"\d+(?:\.\d+)?\s*[*xX]\s*\d+(?:\.\d+)?", s):
        return True
    return False

def validate_workbook_quick(path):
    """
    처리 전 빠른 검증. (bool_ok, 메시지)
    """
    if not path or not os.path.isfile(path):
        return False, "파일이 없습니다."
    low = path.lower()
    if not low.endswith((".xlsx", ".xlsm")):
        return False, "지원 형식은 .xlsx / .xlsm 입니다."
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            if ws is None:
                return False, "활성 시트가 없습니다."
            mr = min(ws.max_row, 10000)
            found = False
            for r in range(14, mr + 1):
                for c in range(13, 19):
                    v = ws.cell(row=r, column=c).value
                    if v is not None and str(v).strip():
                        found = True
                        break
                if found:
                    break
            if not found:
                return False, "행 14 이후 M~R 열에 데이터가 없습니다. 양식을 확인하세요."
        finally:
            wb.close()
        return True, "검증 통과"
    except Exception as e:
        return False, f"엑셀을 열 수 없습니다: {e}"

# UI 키 → (load_database의 category 문자열, 스펙 열 개수)
DB_CATEGORY_SCHEMA = {
    "IC":     ("IC", 1),
    "MOSFET": ("MOSFET", 3),
    "DIODE":  ("RECTIFIER(DIODE)", 3),
    "CAP":    ("CAP", 2),
    "TR":     ("TR", 4),
}

def category_breakdown_text(records):
    """부품(카테고리)별 건수 문자열 생성."""
    from collections import Counter
    order = [
        ("IC", "IC"),
        ("MOSFET", "MOSFET"),
        ("RECTIFIER(DIODE)", "DIODE"),
        ("CAP", "CAP"),
        ("TR", "TR"),
    ]
    cnt = Counter(r["category"] for r in records)
    parts = []
    for key, label in order:
        n = cnt.get(key, 0)
        if n > 0:
            parts.append(f"{label} {n}")
    return " · ".join(parts) if parts else "—"

def find_db_sheet_name(workbook, category_ui_key):
    """카테고리에 해당하는 DB 시트명 반환."""
    sheet_search_key = "DIODE" if category_ui_key == "DIODE" else category_ui_key
    for name in workbook.sheetnames:
        if sheet_search_key in name.upper():
            return name
    return None

def get_category_spec_labels(category_ui_key):
    """
    카테고리별 스펙 라벨을 DB 헤더에서 읽어 반환.
    헤더가 비어 있으면 기본값(스펙 D열 등) 사용.
    """
    if category_ui_key not in DB_CATEGORY_SCHEMA:
        return []
    _, n_spec = DB_CATEGORY_SCHEMA[category_ui_key]
    fallback = [f"스펙 {get_column_letter(4+i)}열" for i in range(n_spec)]
    if not os.path.exists(DB_FILE):
        return fallback

    wb = load_workbook(DB_FILE, read_only=True, data_only=True)
    try:
        sheet_name = find_db_sheet_name(wb, category_ui_key)
        if not sheet_name:
            return fallback
        ws = wb[sheet_name]
        labels = []
        for i in range(n_spec):
            col = 4 + i
            header = ws.cell(row=2, column=col).value
            label = str(header).strip() if header is not None else ""
            if not label:
                label = f"스펙 {get_column_letter(col)}열"
            labels.append(label)
        return labels
    finally:
        wb.close()

def append_part_to_database(category_ui_key, part_name, specs):
    """
    DB/Database.xlsx에 신규 행 추가. C열=부품번호, D열부터 스펙.
    load_database와 동일한 시트 매칭 규칙 사용.
    """
    if category_ui_key not in DB_CATEGORY_SCHEMA:
        raise ValueError("알 수 없는 카테고리입니다.")
    _, n_spec = DB_CATEGORY_SCHEMA[category_ui_key]
    specs = [str(s).strip() for s in specs]
    if len(specs) != n_spec:
        raise ValueError(f"스펙은 {n_spec}개를 입력해야 합니다.")
    if not part_name or not str(part_name).strip():
        raise ValueError("부품명을 입력하세요.")

    wb = load_workbook(DB_FILE, read_only=False, keep_vba=False)
    try:
        sheet_name = find_db_sheet_name(wb, category_ui_key)
        if not sheet_name:
            raise FileNotFoundError(f"카테고리 '{category_ui_key}'에 해당하는 시트를 찾을 수 없습니다.")

        ws = wb[sheet_name]
        pn_norm = str(part_name).strip()
        key_new = normalize_part_key(pn_norm)
        for r in range(3, ws.max_row + 1):
            v = ws.cell(row=r, column=3).value
            if v is None:
                continue
            existing = str(v).strip()
            if existing.upper() == pn_norm.upper():
                raise ValueError(f"동일 부품명이 이미 있습니다: {existing}")
            if key_new and normalize_part_key(existing) == key_new:
                raise ValueError(f"유사 부품명이 이미 있습니다(공백·기호 차이): '{existing}'")

        last = 2
        for r in range(3, ws.max_row + 1):
            v = ws.cell(row=r, column=3).value
            if v is not None and str(v).strip():
                last = r
        new_r = last + 1
        ws.cell(row=new_r, column=3).value = pn_norm
        for i, val in enumerate(specs):
            ws.cell(row=new_r, column=4 + i).value = val
        wb.save(DB_FILE)
    finally:
        wb.close()


def batch_append_parts_to_database(entries, log_func=None):
    """
    여러 부품을 Database.xlsx에 한 번만 열어 추가한다.
    entries: [{"category": "CAP", "part_name": "PN", "specs": ["a","b"]}, ...]
    같은 시트·같은 normalize_part_key 는 건너뛴다(이미 있으면 skipped_dup).
    반환: {"added": int, "skipped_dup": int, "skipped_invalid": int, "errors": [str]}
    """
    log = log_func or (lambda s: None)
    result = {"added": 0, "skipped_dup": 0, "skipped_invalid": 0, "errors": []}
    wb = load_workbook(DB_FILE, read_only=False, keep_vba=False)
    cache = {}
    try:
        for cat in DB_CATEGORY_SCHEMA:
            sn = find_db_sheet_name(wb, cat)
            if not sn:
                continue
            ws = wb[sn]
            keys = set()
            cap_weak = set()
            last = 2
            for r in range(3, ws.max_row + 1):
                v = ws.cell(row=r, column=3).value
                if v is None or not str(v).strip():
                    continue
                c3 = str(v).strip()
                keys.add(normalize_part_key(c3))
                if cat == "CAP":
                    d4 = ws.cell(row=r, column=4).value
                    e5 = ws.cell(row=r, column=5).value
                    cap_weak.add(cap_import_weak_dup_key(c3, [d4, e5]))
                last = r
            cache[cat] = {"ws": ws, "keys": keys, "last": last}
            if cat == "CAP":
                cache[cat]["cap_weak"] = cap_weak

        for ent in entries:
            cat = ent.get("category")
            pn = str(ent.get("part_name") or "").strip()
            specs = ent.get("specs") or []
            if cat not in DB_CATEGORY_SCHEMA:
                result["skipped_invalid"] += 1
                result["errors"].append(f"{pn or '?'}: 알 수 없는 카테고리 {cat}")
                continue
            _, n_spec = DB_CATEGORY_SCHEMA[cat]
            if len(specs) != n_spec:
                result["skipped_invalid"] += 1
                result["errors"].append(f"{pn}: 스펙 개수 오류 (요구 {n_spec}, 실제 {len(specs)})")
                continue
            if not pn:
                result["skipped_invalid"] += 1
                continue
            if cat not in cache:
                result["errors"].append(f"{pn}: 시트 없음 — {cat}")
                continue
            st = cache[cat]
            key_new = normalize_part_key(pn)
            if key_new in st["keys"]:
                result["skipped_dup"] += 1
                continue
            if cat == "CAP" and st.get("cap_weak") is not None:
                wk = cap_import_weak_dup_key(pn, specs)
                if cap_part_row_is_incomplete(pn) and wk in st["cap_weak"]:
                    result["skipped_dup"] += 1
                    continue
            new_r = st["last"] + 1
            st["ws"].cell(row=new_r, column=3).value = pn
            for i, val in enumerate(specs):
                st["ws"].cell(row=new_r, column=4 + i).value = str(val).strip() if val is not None else ""
            st["keys"].add(key_new)
            if cat == "CAP" and st.get("cap_weak") is not None:
                st["cap_weak"].add(cap_import_weak_dup_key(pn, specs))
            st["last"] = new_r
            result["added"] += 1

        wb.save(DB_FILE)
        log(f"[batch_append] 저장 완료: 추가 {result['added']}, 이미있음 {result['skipped_dup']}, 무효 {result['skipped_invalid']}")
    except Exception as e:
        result["errors"].append(str(e))
        log(f"[batch_append] 오류: {e}")
        raise
    finally:
        wb.close()
    return result


# ─────────────────────────────────────────────
# Database 로드 (ZIP/XML 직접 파싱 - 속도 및 정확도 최대화)
# ─────────────────────────────────────────────
def load_database():
    import zipfile
    import xml.etree.ElementTree as ET

    if not os.path.exists(DB_FILE):
        raise FileNotFoundError(f"Database 파일을 찾을 수 없습니다:\n{DB_FILE}")

    records = []
    
    def find_cell_index(cell_ref):
        col_str = "".join(filter(str.isalpha, cell_ref))
        idx = 0
        for char in col_str:
            idx = idx * 26 + (ord(char.upper()) - ord('A') + 1)
        return idx - 1

    try:
        with zipfile.ZipFile(DB_FILE, 'r') as zf:
            # 1. 공통 문자열 로드
            strings = []
            try:
                with zf.open('xl/sharedStrings.xml') as f:
                    root = ET.parse(f).getroot()
                    ns = {'m': root.tag.split('}')[0].strip('{')}
                    for t in root.findall('.//m:t', ns):
                        strings.append(t.text if t.text else "")
            except: pass

            # 2. 시트 리스트 로드
            sheet_paths = {}
            rid_map = {}
            try:
                with zf.open('xl/_rels/workbook.xml.rels') as f:
                    root = ET.parse(f).getroot()
                    ns = {'m': root.tag.split('}')[0].strip('{')}
                    for r in root.findall('.//m:Relationship', ns):
                        rid_map[r.attrib.get('Id')] = r.attrib.get('Target')
                
                with zf.open('xl/workbook.xml') as f:
                    root = ET.parse(f).getroot()
                    ns = {'m': root.tag.split('}')[0].strip('{')}
                    for s in root.findall('.//m:sheet', ns):
                        name = s.attrib.get('name', '').upper().strip()
                        rid = s.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                        path = rid_map.get(rid)
                        if path:
                            if not path.startswith('worksheets/'): path = f"worksheets/{os.path.basename(path)}"
                            sheet_paths[name] = f"xl/{path}"
            except: pass

            # 모든 시트: C열(index 2)이 부품번호, D열(index 3)부터가 스펙
            categories = {
                "IC":     ["IC", 1],      # D
                "MOSFET": ["MOSFET", 3],  # D, E, F
                "DIODE":  ["RECTIFIER(DIODE)", 3], # D, E, F
                "CAP":    ["CAP", 2],     # D, E
                "TR":     ["TR", 4]       # D, E, F, G
            }

            for key, (cat_name, spec_count) in categories.items():
                xpath = None
                for s_name in sheet_paths:
                    if key in s_name:
                        xpath = sheet_paths[s_name]
                        break
                
                if not xpath: continue
                
                try:
                    with zf.open(xpath) as f:
                        root = ET.parse(f).getroot()
                        ns = {'m': root.tag.split('}')[0].strip('{')}
                        count = 0
                        for row_node in root.findall('.//m:row', ns):
                            r_idx = int(row_node.attrib.get('r', 0))
                            if r_idx < 3: continue 
                            
                            row_vals = {}
                            for c in row_node.findall('m:c', ns):
                                ref = c.attrib.get('r')
                                idx = find_cell_index(ref)
                                v = c.find('m:v', ns)
                                val = v.text if v is not None else ""
                                if c.attrib.get('t') == 's' and val.isdigit():
                                    try: val = strings[int(val)]
                                    except: pass
                                row_vals[idx] = val
                            
                            # C열(index 2)이 부품번호!!
                            pn = row_vals.get(2)
                            if pn and str(pn).strip() and str(pn).strip().upper() not in ["IC", "MOSFET", "DIODE", "CAP", "TR", "PART NUMBER"]:
                                # D열(index 3)부터 스펙 시작
                                specs = [str(row_vals.get(i, "")).strip() for i in range(3, 3 + spec_count)]
                                records.append({
                                    "category": cat_name,
                                    "part_number": str(pn).strip(),
                                    "specs": specs
                                })
                                count += 1
                        print(f"[DB 로드] {key} 연관 시트: {count}개 로드 완료")
                except: pass

    except Exception as e:
        print(f"[오류] DB ZIP 파싱 중 실패: {e}")

    return records


# ─────────────────────────────────────────────
# Spec 숫자 파싱
# ─────────────────────────────────────────────
def parse_spec_numbers(spec_str):
    clean = str(spec_str).replace('\n', '/').replace('\r', '')
    parts = clean.split('/')
    numbers = []
    
    pattern = re.compile(r'(\d+(?:\.\d+)?)\s*([a-zA-Z]+)?')
    
    for part in parts:
        m = pattern.search(part.strip())
        if m:
            val = float(m.group(1))
            unit = m.group(2)
            if unit:
                unit = unit.strip().lower()
                if unit == 'mv':
                    val = val / 1000.0
                    unit = 'v'
                elif unit == 'v':
                    unit = 'v'
                elif unit == 'ma':
                    val = val / 1000.0
                    unit = 'a'
                elif unit == 'a':
                    unit = 'a'
                else:
                    pass
            numbers.append((val, unit))
    return numbers


# ─────────────────────────────────────────────
# 측정값 색상 처리
# ─────────────────────────────────────────────
def apply_measurement_color(ws, match_row, category, spec_str, log_func):
    meas_row   = match_row + 6
    col_start  = 28
    col_end    = 91

    spec_limits = parse_spec_numbers(spec_str)
    if not spec_limits:
        log_func(f"  [색상 건너뜀] spec 숫자 없음: {spec_str}")
        return

    log_func(f"  [색상 검사] 행{meas_row} {get_column_letter(col_start)}:{get_column_letter(col_end)} "
             f"| spec 한계(정규화)={spec_limits}")

    red_count = 0
    font_red   = InlineFont(color='FF0000')
    font_black = InlineFont(color='000000')

    pattern = re.compile(r'(\d+(?:\.\d+)?)\s*([a-zA-Z]+)?')

    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=meas_row, column=col)
        val_str = str(cell.value) if cell.value is not None else ""
        if not val_str.strip() or val_str == "None":
            continue

        parts = val_str.split('/')
        has_red = False
        rich_blocks = []

        for i, part in enumerate(parts):
            m = pattern.search(part)
            font_to_use = font_black
            
            if m:
                meas_val = float(m.group(1))
                unit = m.group(2)
                
                if unit:
                    unit = unit.strip().lower()
                    if unit == 'mv':
                        meas_val = meas_val / 1000.0
                    elif unit == 'ma':
                        meas_val = meas_val / 1000.0
                
                limit_val, _ = spec_limits[i % len(spec_limits)]
                
                if meas_val > limit_val:
                    font_to_use = font_red
                    has_red = True
            
            rich_blocks.append(TextBlock(font_to_use, part))
            
            if i < len(parts) - 1:
                rich_blocks.append(TextBlock(font_black, "/"))

        if has_red:
            cell.value = CellRichText(*rich_blocks)
            red_count += 1
            log_func(f"  [빨간색] {get_column_letter(col)}{meas_row} '{val_str}' -> 부분 빨간색 적용", color="red")

    if red_count:
        log_func(f"  -> {red_count}개 셀 빨간색 처리 완료", color="yellow")
    else:
        log_func(f"  -> 한계 초과 없음")


# ─────────────────────────────────────────────
# 엑셀1 처리
# ─────────────────────────────────────────────
def backup_excel(excel_path, log_func):
    backup_dir = os.path.join(BASE_DIR, "backup")
    os.makedirs(backup_dir, exist_ok=True)

    fname      = os.path.basename(excel_path)
    name, ext  = os.path.splitext(fname)
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{name}_{timestamp}{ext}"
    backup_path = os.path.join(backup_dir, backup_name)

    shutil.copy2(excel_path, backup_path)
    log_func(f"[백업] backup\\{backup_name} 저장 완료")
    return backup_path


def _col_width_to_pixels(ws, col_idx: int) -> int:
    """openpyxl 컬럼 너비(문자 단위)를 대략 픽셀로 변환."""
    letter = get_column_letter(col_idx)
    w = ws.column_dimensions[letter].width
    if w is None:
        w = 8.43  # Excel default
    return int(w * 7 + 5)


def _row_height_to_pixels(ws, row_idx: int) -> int:
    """openpyxl 행 높이(pt)를 픽셀로 변환."""
    h = ws.row_dimensions[row_idx].height
    if h is None:
        h = 15  # Excel default in points
    return int(points_to_pixels(h))


def get_range_pixel_size(ws, target_range: str):
    """
    예: 'AB14:CM20' 범위의 픽셀 너비/높이 계산.
    """
    min_col, min_row, max_col, max_row = range_boundaries(target_range)
    width_px = sum(_col_width_to_pixels(ws, c) for c in range(min_col, max_col + 1))
    height_px = sum(_row_height_to_pixels(ws, r) for r in range(min_row, max_row + 1))
    anchor = f"{get_column_letter(min_col)}{min_row}"
    return width_px, height_px, anchor


def add_image_fit_to_range(ws, image_path: str, target_range: str, mode: str = "contain", padding_px: int = 2):
    """
    지정 범위 크기에 맞춰 이미지 자동 리사이즈 후 삽입.
    - mode='contain': 비율 유지, 잘림 없이 범위 안에 맞춤
    - mode='cover': 비율 유지, 범위를 꽉 채움(일부 잘릴 수 있음)
    """
    img = XLImage(image_path)
    tw, th, anchor = get_range_pixel_size(ws, target_range)
    tw = max(1, tw - padding_px * 2)
    th = max(1, th - padding_px * 2)
    iw, ih = max(1, int(img.width)), max(1, int(img.height))
    if mode == "cover":
        scale = max(tw / iw, th / ih)
    else:
        scale = min(tw / iw, th / ih)
    img.width = max(1, int(iw * scale))
    img.height = max(1, int(ih * scale))
    img.anchor = anchor
    ws.add_image(img)
    return img.width, img.height, anchor


def _norm_text(s: str) -> str:
    return re.sub(r"[\s\-_]+", "", str(s or "")).upper()


def _is_schematic_ref_token(tok: str) -> bool:
    """로케이션(RefDes)처럼 보이는 짧은 토큰 (예: CP815S, QM801CS, ICM801S)."""
    t = str(tok or "").strip()
    if not t or len(t) > 22 or len(t) < 3:
        return False
    return bool(re.match(r"^[A-Z]{1,6}\d{1,5}[A-Z0-9]{0,4}$", t, re.I))


# Stress Rule 시트 1열에 같이 실리는 양식/헤더용 키 — 실제 부품 Function이 아님
_RULE_SHEET_TEMPLATE_FUNC_KEYS = frozenset(
    {
        "FUNCTION",
        "RESULT",
        "SPEC",
        "PARTNUMBER",
        "PARTNUBMER",
        "LOCNUMPARTNUMBER",
        "LOCNUMPARTNUBMER",
    }
)

_MPN_LIKE_TOKEN = re.compile(r"^[A-Z][A-Z0-9\-]{7,}$", re.I)


def _strip_mpn_tokens_for_display(text: str) -> str:
    """제조사 MPN(긴 알파넘)은 뺀 뒤 PL 스펙 위주 문자열만 남긴다."""
    s = str(text or "").strip()
    if not s:
        return s
    tokens = re.split(r"\s+", s)
    kept = [t for t in tokens if t and not _MPN_LIKE_TOKEN.match(t)]
    out = " ".join(kept).strip()
    return out if out else s


def _mr_row_has_schematic_ref(m_values: list) -> bool:
    """M~R 중 로케이션(RefDes) 토큰이 하나라도 있을 때만 미일치 신규 후보로 친다."""
    for m in m_values:
        for w in re.split(r"[\s/\n\r]+", str(m)):
            w = w.strip(" ,.;")
            if _is_schematic_ref_token(w):
                return True
    return False


def _is_mr_part_sheet_header_row(m_values: list, joined: str) -> bool:
    """Loc. Num. Part Number 등 M~R 헤더 행이면 신규 후보에서 제외."""
    blob = _norm_text(joined)
    if not blob:
        return False
    if "LOC" in blob and "NUM" in blob and ("PARTNUB" in blob or "PARTNUM" in blob):
        return True
    if "LOCNUMPART" in blob:
        return True
    if "PARTNUBMER" in blob:
        return True
    if len(m_values) == 1:
        only = _norm_text(str(m_values[0]))
        if "LOCNUM" in only and "PART" in only:
            return True
    return False


def _resolve_rule_func_key_for_cell(tk: str, rule_applicable: dict) -> str | None:
    """G~L 한 칸에서 rule_applicable에 대응하는 키를 찌른다. 양식용 키(FUNCTION 등)는 무시."""
    if not tk or not rule_applicable:
        return None
    if tk in rule_applicable and tk not in _RULE_SHEET_TEMPLATE_FUNC_KEYS:
        return tk
    for rk in rule_applicable.keys():
        if not rk or rk in _RULE_SHEET_TEMPLATE_FUNC_KEYS:
            continue
        if rk in tk or tk in rk:
            return rk
    return None


def _unmatched_group_key_and_display(pn_str: str) -> tuple[str, str]:
    """
    미일치 검색어 문자열을 '실제 부품·스펙' 기준으로 묶기 위한 (키, 표시문구).
    '로케이션 / 품목·스펙' 형태면 앞의 로케이션은 그룹 키에서 제외해 동일 품목을 1건으로 합친다.
    """
    s = str(pn_str or "").strip()
    if not s:
        return ("", "")
    s_one = re.sub(r"[\r\n]+", " ", s)
    segs = [p.strip() for p in re.split(r"\s*/\s*", s_one) if p.strip()]
    if len(segs) >= 2 and _is_schematic_ref_token(segs[0]):
        identity = " / ".join(segs[1:]).strip()
        if identity:
            identity = _strip_mpn_tokens_for_display(identity)
            disp = identity if len(identity) <= 180 else identity[:177] + "..."
            return (_norm_text(identity), disp)
    if len(segs) == 1:
        words = segs[0].split()
        if len(words) >= 2 and _is_schematic_ref_token(words[0]):
            identity = " ".join(words[1:]).strip()
            if identity:
                identity = _strip_mpn_tokens_for_display(identity)
                disp = identity if len(identity) <= 180 else identity[:177] + "..."
                return (_norm_text(identity), disp)
    s_disp = _strip_mpn_tokens_for_display(s_one)
    disp = s_disp if len(s_disp) <= 180 else s_disp[:177] + "..."
    return (_norm_text(s_disp), disp)


def _extract_location_label_from_unmatched(pn_str: str) -> str | None:
    """미일치 문자열에서 로케이션(RefDes)만 뽑는다. 행 번호 대신 표시용."""
    s = str(pn_str or "").strip()
    if not s:
        return None
    s_one = re.sub(r"[\r\n]+", " ", s)
    segs = [p.strip() for p in re.split(r"\s*/\s*", s_one) if p.strip()]
    if not segs:
        return None
    if _is_schematic_ref_token(segs[0]):
        return segs[0]
    if len(segs) == 1:
        words = segs[0].split()
        if words and _is_schematic_ref_token(words[0]):
            return words[0]
    return None


def _rule_type_key(type_text: str) -> str:
    t = str(type_text or "").upper()
    if "WORST" in t:
        # 90~264/264~90의 worst transient 계열
        return "WORST"
    if "NORMAL" in t:
        return "NORMAL"
    if "TURN-ON" in t or "TURN ON" in t:
        return "TURN_ON"
    if "TURN-OFF" in t or "TURN OFF" in t:
        return "TURN_OFF"
    # fallback: 검색되는 키워드 기반
    if "TURN" in t and "ON" in t:
        return "TURN_ON"
    if "TURN" in t and "OFF" in t:
        return "TURN_OFF"
    if "TRANSIENT" in t:
        return "TRANSIENT"
    return _norm_text(t)[:12]


def _rule_group_key(group_text: str) -> str:
    g = str(group_text or "").upper()
    g = g.replace(" ", "")
    # 90Vac / 264Vac / 90~264Vac / 264~90Vac
    if "90VAC" in g and "264" not in g:
        return "90VAC"
    if "264VAC" in g and "90" not in g:
        return "264VAC"
    # 90 ~ 264Vac
    # (중요) 설명 문장처럼 90V/264V가 들어있는 경우는 제외하려고 VAC 문자열이 같이 있을 때만 그룹으로 판단
    if "90" in g and "264" in g and "VAC" in g:
        # 방향성(앞에 나온 숫자 기준)
        i90 = g.find("90")
        i264 = g.find("264")
        return "90_264VAC" if i90 < i264 else "264_90VAC"
    return _norm_text(g)[:16]


def _load_stress_rule_index(rule_xlsx: str, log_func=None):
    from openpyxl import load_workbook as _lb

    # merged cell 값을 읽어야 하므로 read_only=False로 로드
    wb = _lb(rule_xlsx, read_only=False, data_only=True)
    ws = wb.active  # Sheet1

    # rule 함수 행: col1이 채워진 행부터 끝까지
    rule_func_rows = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None or str(v).strip() == "":
            continue
        rule_func_rows.append(r)

    # 조건 헤더 열: row2에 값이 있는 컬럼
    cond_cols = []
    for c in range(2, ws.max_column + 1):
        tv = ws.cell(row=2, column=c).value
        if tv is None or str(tv).strip() == "":
            continue
        cond_cols.append(c)

    cond_key_by_rule_col = {}
    for c in cond_cols:
        group_text = ws.cell(row=1, column=c).value
        # 병합셀로 인해 (row=1, col=c) 위치엔 값이 None으로 남는 경우가 많다.
        if group_text is None:
            for mrange in ws.merged_cells.ranges:
                if mrange.min_row <= 1 <= mrange.max_row and mrange.min_col <= c <= mrange.max_col:
                    group_text = ws.cell(row=mrange.min_row, column=mrange.min_col).value
                    break
        type_text = ws.cell(row=2, column=c).value
        cond_key_by_rule_col[c] = f"{_rule_group_key(group_text)}:{_rule_type_key(type_text)}"

    rule_applicable = {}
    for r in rule_func_rows:
        func = ws.cell(row=r, column=1).value
        func_key = _norm_text(func)
        rule_applicable.setdefault(func_key, {})
        for c in cond_cols:
            val = ws.cell(row=r, column=c).value
            cond_key = cond_key_by_rule_col[c]
            rule_applicable[func_key][cond_key] = not (val is None or str(val).strip() == "")

    try:
        wb.close()
    except Exception:
        pass

    if log_func:
        log_func(f"[룰] Stress Rule loaded: funcs={len(rule_applicable)}, conds={len(cond_cols)}")
    return rule_applicable, list(cond_key_by_rule_col.values())


def process_excel(excel_path, db_records, log_func, options=None):
    opts = options or {}
    do_norm = opts.get("normalize_cell_text", True)
    dry_run = opts.get("dry_run", False)
    progress_cb = opts.get("progress_cb")
    src_name = os.path.basename(excel_path)

    if dry_run:
        log_func(f"[드라이런] {src_name} — 원본 백업·Result 저장 생략 (매칭만 시뮬레이션)")
    else:
        backup_excel(excel_path, log_func)

    wb = load_workbook(excel_path)
    ws = wb.active

    def _set_value_safely(row_idx, col_idx, value):
        """
        병합셀 좌표면 병합영역 좌상단 셀에 기록한다.
        """
        target = f"{get_column_letter(col_idx)}{row_idx}"
        for mrange in ws.merged_cells.ranges:
            if target in mrange:
                ws.cell(row=mrange.min_row, column=mrange.min_col).value = value
                return (mrange.min_row, mrange.min_col, True)
        ws.cell(row=row_idx, column=col_idx).value = value
        return (row_idx, col_idx, False)

    # X(비활성) 표시용 셀 스타일: "X 텍스트만"이 아니라 셀 자체를 그을리게 보이도록 처리
    _disabled_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _disabled_font = Font(color="7F7F7F", strike=True)
    _disabled_alignment = Alignment(horizontal="center", vertical="center")
    _disabled_border_side = Side(style="thin", color="BFBFBF")
    _disabled_border = Border(
        left=_disabled_border_side,
        right=_disabled_border_side,
        top=_disabled_border_side,
        bottom=_disabled_border_side,
    )

    def _apply_disabled_style(row_idx, col_idx):
        addr = f"{get_column_letter(col_idx)}{row_idx}"
        # 병합된 범위면 범위 전체에 스타일을 적용(비활성처럼 보이게)
        for mrange in ws.merged_cells.ranges:
            if addr in mrange:
                for rr in range(mrange.min_row, mrange.max_row + 1):
                    for cc in range(mrange.min_col, mrange.max_col + 1):
                        cell = ws.cell(row=rr, column=cc)
                        cell.fill = _disabled_fill
                        cell.font = _disabled_font
                        cell.alignment = _disabled_alignment
                        cell.border = _disabled_border
                return

        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = _disabled_fill
        cell.font = _disabled_font
        cell.alignment = _disabled_alignment
        cell.border = _disabled_border

    def _detect_remote_start_row():
        # Remote On/Off(Stand By Block) 구역은 X 마킹 대상 제외.
        # 단, "Standby mode" 같은 다른 문구 때문에 오탐되지 않도록
        # 반드시 "REMOTE" + "ON/OFF" 조합이 포함된 헤더 셀만 사용한다.
        max_r = min(ws.max_row or 1, 2000)
        # Remote 헤더는 보통 왼쪽(대략 A~D)에 위치하므로 컬럼 범위를 좁혀 오탐을 줄인다.
        max_c = min(ws.max_column or 1, 10)
        try:
            for r in range(1, max_r + 1):
                for c in range(1, max_c + 1):
                    v = ws.cell(row=r, column=c).value
                    if not v:
                        continue
                    t = str(v).upper().replace(" ", "")
                    if "REMOTE" in t and ("ON/OFF" in t or "ONOFF" in t or ("ON" in t and "OFF" in t)):
                        return r
            return None
        except Exception:
            return None

    def _detect_template_cond_col_mapping(expected_cond_keys):
        # 템플릿에서 rule의 cond_key에 해당하는 컬럼을 찾는다.
        # cond_key 형식: {group}:{type} (예: 90VAC:NORMAL)
        header_row_scan_max = 20
        col_scan_max = min(ws.max_column or 1, 250)

        group_by_col = {}
        # merged 범위 기반 group 매핑
        try:
            for rng in ws.merged_cells.ranges:
                if rng.min_row <= header_row_scan_max and rng.max_row <= header_row_scan_max:
                    v = ws.cell(row=rng.min_row, column=rng.min_col).value
                    if not v:
                        continue
                    g_key = _rule_group_key(v)
                    if g_key in ("90VAC", "264VAC", "90_264VAC", "264_90VAC"):
                        for c in range(rng.min_col, rng.max_col + 1):
                            group_by_col[c] = g_key
        except Exception:
            pass

        # 비-merged도 보완
        for c in range(1, col_scan_max + 1):
            if c in group_by_col:
                continue
            for r in range(1, header_row_scan_max + 1):
                v = ws.cell(row=r, column=c).value
                if not v:
                    continue
                g_key = _rule_group_key(v)
                if g_key in ("90VAC", "264VAC", "90_264VAC", "264_90VAC"):
                    group_by_col[c] = g_key
                    break

        type_by_col = {}
        for c in range(1, col_scan_max + 1):
            for r in range(2, header_row_scan_max + 1):
                v = ws.cell(row=r, column=c).value
                if not v:
                    continue
                t = str(v).upper()
                if ("NORMAL" in t) or ("TURN" in t and ("ON" in t or "OFF" in t)) or ("WORST" in t) or ("TRANSIENT" in t):
                    type_by_col[c] = _rule_type_key(v)
                    break

        cond_col_map = {}
        for c, g_key in group_by_col.items():
            if c not in type_by_col:
                continue
            t_key = type_by_col[c]
            cond_key = f"{g_key}:{t_key}"
            if cond_key in expected_cond_keys:
                cond_col_map[cond_key] = c
        return cond_col_map

    matched           = 0
    checked           = 0
    unmatched         = []
    row_step          = 7
    row               = 14

    # Stress Rule 기반 X 마킹 사전 준비 (Stress Analysis 구역만)
    rule_applicable = {}
    cond_cols_map = {}
    remote_start_row = None
    remote_block_start_row = None
    try:
        if os.path.exists(STRESS_RULE_FILE):
            rule_applicable, rule_cond_key_list = _load_stress_rule_index(STRESS_RULE_FILE, log_func=log_func)
            cond_cols_map = _detect_template_cond_col_mapping(set(rule_cond_key_list))
        remote_start_row = _detect_remote_start_row()
        if remote_start_row is not None and remote_start_row >= 14:
            # block 시작 행(14,21,28,...) 기준으로 스킵
            remote_block_start_row = 14 + 7 * ((remote_start_row - 14) // 7)
        if not cond_cols_map:
            log_func("[X] Stress Rule cond 컬럼 매핑 실패(헤더 구조가 다를 수 있음) - X 마킹 건너뜀", color="yellow")
        else:
            log_func(
                f"[X] Stress Rule 기반 X 마킹 준비 완료 - cond_cols={len(cond_cols_map)}, "
                f"remote_start={remote_start_row}, remote_block_start={remote_block_start_row}",
                color="green",
            )
    except Exception as e:
        log_func(f"[X] Stress Rule 준비 실패: {e}", color="yellow")

    def find_real_max_row(ws, start_row=14):
        # 엑셀의 가짜 max_row(100만 행) 문제를 해결하기 위해 
        # 상단부터 10000행까지만 데이터가 있는지 빠르게 확인
        last_found = start_row
        max_to_check = min(ws.max_row, 10000) # 현실적으로 10000행 이상 데이터가 있지는 않음
        
        # iter_rows가 ws.cell()보다 수천배 빠릅니다.
        for r_idx, row_cells in enumerate(ws.iter_rows(min_row=start_row, max_row=max_to_check, min_col=13, max_col=18, values_only=True), start=start_row):
            if any(cell for cell in row_cells):
                last_found = r_idx
        return last_found

    print(f"[알림] 엑셀 파일 분석 중... 잠시만 기다려 주세요.")
    real_max = find_real_max_row(ws)
    log_func(f"[진행] 실제 데이터 탐지 완료: 행 {real_max}까지 처리를 시작합니다.")

    # 양식이 7행 블록 구조이므로, "블록 시작행만"이 아니라
    # 블록 내부에서도 M~R(13~18)에 실제 데이터가 있는 행만 검사한다.
    # 특히 Remote On/Off 섹션은 블록 시작행이 아니라 내부 행에 부품명이 들어있는 경우가 있어 누락이 발생한다.
    rows_to_check = []
    seen = set()
    block_starts = list(range(14, real_max + 1, row_step))
    for bs in block_starts:
        for r in range(bs, min(bs + row_step, real_max + 1)):
            has_m = False
            for c in range(13, 19):  # M~R
                if ws.cell(row=r, column=c).value:
                    has_m = True
                    break
            if has_m and r not in seen:
                seen.add(r)
                rows_to_check.append(r)

    if not rows_to_check:
        rows_to_check = list(range(14, real_max + 1, row_step))
    total_rows = max(1, len(rows_to_check))
    if callable(progress_cb):
        progress_cb(0, total_rows, f"{src_name} 시작")

    for row_idx, row in enumerate(rows_to_check, start=1):
        if callable(progress_cb):
            progress_cb(row_idx, total_rows, f"{src_name} 처리 중")

        # "function 있는 칸만" 신규 부품 카운팅에 반영하기 위해,
        # 현재 row의 G~L(Function 열)에서 실제 Function 키를 찾는다.
        # (템플릿의 헤더/설명 문구 때문에 M~R에 글이 있어도 Function이 없는 행은 제외)
        func_key_for_count = None
        if rule_applicable:
            for col_idx in range(7, 13):  # G~L
                v = ws.cell(row=row, column=col_idx).value
                if not v:
                    continue
                tk = _norm_text(v)
                fk = _resolve_rule_func_key_for_cell(tk, rule_applicable)
                if fk:
                    func_key_for_count = fk
                    break
        else:
            for col_idx in range(7, 13):
                v = ws.cell(row=row, column=col_idx).value
                if not v:
                    continue
                t = str(v).strip().upper()
                raw_norm = _norm_text(v)
                if (
                    t
                    and raw_norm not in _RULE_SHEET_TEMPLATE_FUNC_KEYS
                    and t not in ("FUNCTION", "RESULT", "SPEC.")
                ):
                    func_key_for_count = raw_norm
                    break

        # Stress Analysis 구역에서만 파형 슬롯/결과값 비활성 칸 X 마킹(텍스트는 비움)
        if rule_applicable and cond_cols_map:
            # 해당 row가 속한 7행 블록의 시작행을 기준으로 "슬롯 행/결과 행"을 맞춘다.
            block_start_row = row - ((row - 14) % row_step) if row >= 14 else row
            slot_row = block_start_row
            result_row = block_start_row + (row_step - 1)
            if remote_block_start_row is not None and slot_row >= remote_block_start_row:
                slot_row = None
            if slot_row is None:
                # Remote 구역(Standby/On-Off)로 판정되는 블록은 X 마킹/비활성 처리 제외
                is_remote_block = True
            else:
                is_remote_block = False
            # Remote On/Off / Stand By Block 구역은 X 마킹에서 제외
            func_key_found = None
            if not is_remote_block:
                for col_idx in range(7, 13):  # G~L
                    v = ws.cell(row=row, column=col_idx).value
                    if not v:
                        continue
                    t_raw = str(v).upper().replace(" ", "")
                    if ("REMOTE" in t_raw and ("ON" in t_raw or "OFF" in t_raw)) or (
                        "STAND" in t_raw and "BY" in t_raw
                    ):
                        is_remote_block = True
                        break

                    tk = _norm_text(v)
                    func_key_found = _resolve_rule_func_key_for_cell(tk, rule_applicable)
                    if func_key_found:
                        break

                if (not is_remote_block) and func_key_found:
                    for cond_key, cond_col in cond_cols_map.items():
                        applicable = rule_applicable.get(func_key_found, {}).get(cond_key, True)
                        if not applicable:
                            # "X" 텍스트는 제거하고(값 비움) 회색 비활성 스타일만 남긴다.
                            # 슬롯 행(slot_row) + 결과 행(result_row)까지 같이 비활성 처리
                            _set_value_safely(slot_row, cond_col, "")
                            _apply_disabled_style(slot_row, cond_col)
                            _set_value_safely(result_row, cond_col, "")
                            _apply_disabled_style(result_row, cond_col)
        # 진행 상황 표시 (20행 단위)
        if (row - 14) % 21 == 0:
            print(f"[진행] 행 {row} 처리 중...")
        
        # 디버그용 DB 레코드 덤프 (최초 1회만)
        if row == 14:
            try:
                with open("db_debug_dump.txt", "w", encoding="utf-8") as f:
                    for r in db_records:
                        f.write(f"{r}\n")
                log_func("[시스템] DB 덤프 완료 (db_debug_dump.txt 확인 가능)")
            except: pass

        m_values = []
        for col_idx in range(13, 19): # M, N, O, P, Q, R (13~18)
            val = ws.cell(row=row, column=col_idx).value
            if val:
                t = normalize_cell_text(val) if do_norm else str(val).strip()
                if t:
                    m_values.append(t)

        use_pl = opts.get("use_part_list_lookup", True) and HAS_PART_LIST
        pl_lookup = opts.get("pl_lookup") or {}
        vendor_rank_by_loc = opts.get("pl_vendor_rank_by_loc") or {}
        if use_pl and pl_lookup:
            m_values = enrich_m_values_from_pl(m_values, pl_lookup, vendor_rank_by_loc, log_func)

        if not m_values:
            continue

        db_match = None
        current_m_str = " / ".join(m_values) 
        
        # 디버그: 현재 행에서 찾은 문자열들 출력
        # log_func(f"[디버그] 행{row} 검색어: {m_values}")

        for rec in db_records:
            pn = rec["part_number"]
            if not pn: continue
            
            match_found = False
            for m_str in m_values:
                # 1. 길이 4자 미만은 정확히 일치해야 함 (예: 'IC'가 'ICM801'에 매칭되는 것 방지)
                if len(pn) < 4:
                    if pn.lower() == m_str.lower():
                        match_found = True
                        break
                    # 셀 내부 단어와 정확히 일치하는지 확인
                    words = m_str.replace('\n', ' ').replace('\r', ' ').replace('/', ' ').split()
                    if any(pn.lower() == w.strip().lower() for w in words):
                        match_found = True
                        break
                else:
                    # 4자 이상은 기존처럼 포함 여부 확인
                    if pn.lower() in m_str.lower():
                        match_found = True
                        break
                    
                    # 줄바꿈이나 공백, 특수기호로 쪼개서 개별 단어와 일치하는지 확인
                    clean_str = m_str.replace('\n', ' ').replace('\r', ' ').replace('/', ' ').replace('(', ' ').replace(')', ' ')
                    words = clean_str.split()
                    if any(pn.lower() == w.strip().lower() for w in words):
                        match_found = True
                        break
            
            if match_found:
                db_match = rec
                break

        if db_match:
            checked += 1
            specs = db_match["specs"]
            valid_specs = [str(s).strip() for s in specs if s is not None and str(s).strip() != ""]
            combined_spec = " / ".join(valid_specs)
            
            # Loc. Num/Part Number 영역(M~R)에 부품명을 덧붙이고, V열(22번)에 스펙 입력
            appended = False
            for col_idx in range(13, 19):  # M~R
                cell_val = ws.cell(row=row, column=col_idx).value
                if not cell_val:
                    continue
                origin = str(cell_val).strip()
                if not origin:
                    continue
                low_origin = origin.lower()
                low_pn = str(db_match["part_number"]).strip().lower()
                if low_pn in low_origin:
                    appended = True
                    break
                ws.cell(row=row, column=col_idx).value = f"{origin}\n{db_match['part_number']}"
                appended = True
                break

            if not appended:
                # M~R이 비어있는 예외 케이스면 M열에 부품명 기록
                ws.cell(row=row, column=13).value = db_match["part_number"]
                log_func(f"[안내] 행{row} M~R 비어 있어 M열에 부품명 기록", color="yellow")

            sp_r, sp_c, sp_merged = _set_value_safely(row, 22, combined_spec)
            if sp_merged:
                sp_pos = f"{get_column_letter(sp_c)}{sp_r}"
                log_func(f"[안내] 행{row} 스펙 병합셀 기록: {sp_pos}", color="yellow")
            
            log_func(
                f"[{db_match['category']}] 행{row} 매칭성공: '{db_match['part_number']}' "
                f"-> 입력스펙: {combined_spec}", color="blue"
            )
            matched += 1
            apply_measurement_color(ws, row, db_match["category"], combined_spec, log_func)
        else:
            # 특정 부품(예: MBRF2080CTP)이 안 나오는 이유를 찾기 위한 임시 디버그
            if any("MBRF2080CTP" in s.upper() for s in m_values):
                log_func(f"[디버그] 행{row}에서 MBRF2080CTP 발견했으나 DB 매칭 실패!!", color="yellow")
            
            # Function(실제 룰 키) + M~R에 Ref 로케이션이 있을 때만 신규 후보에 포함
            if (
                func_key_for_count
                and _mr_row_has_schematic_ref(m_values)
                and not _is_mr_part_sheet_header_row(m_values, current_m_str)
            ):
                unmatched.append((src_name, row, current_m_str))

    result_path = None
    if dry_run:
        wb.close()
        log_func(f"[드라이런] {src_name} 처리 끝 — 파일 저장 없음")
    else:
        result_dir = os.path.join(BASE_DIR, "Result")
        os.makedirs(result_dir, exist_ok=True)
        fname = os.path.basename(excel_path)
        name, ext = os.path.splitext(fname)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        result_name = f"{name}_result_{timestamp}{ext}"
        result_path = os.path.join(result_dir, result_name)
        wb.save(result_path)
        wb.close()
        log_func(f"[저장] Result\\{result_name}")
    
    unmatched_count = len(unmatched)
    log_func(f"\n완료: 총 {checked}개 확인, {matched}개 일치, {unmatched_count}개 미일치")
    if callable(progress_cb):
        progress_cb(total_rows, total_rows, f"{src_name} 완료")
    
    return matched, checked, result_path, unmatched

# ─────────────────────────────────────────────
# v0.dev 스타일 모던 UI 클래스 컨ポー넌트
# ─────────────────────────────────────────────

# ─────────────────────────────────────────────
# Enterprise Light Theme (Professional Blue-Gray)
# ─────────────────────────────────────────────
BG_MAIN      = "#f8fafc"   # Slate 50 - Clean base
BG_CARD      = "#ffffff"   # Pure white cards
BG_HEADER    = "#f1f5f9"   # Slate 100 - Subtle header bar
BORDER       = "#e2e8f0"   # Slate 200 - Soft borders
BORDER_FOCUS = "#94a3b8"   # Slate 400 - Focus state
SHADOW       = "#cbd5e1"   # Slate 300 - Card shadow tint
PRIMARY      = "#2563eb"   # Blue 600 - Brand accent
PRIMARY_HOV  = "#1d4ed8"   # Blue 700 - Hover
PRIMARY_SOFT = "#eff6ff"   # Blue 50 - Soft highlight
TEXT_PRI     = "#0f172a"   # Slate 900
TEXT_SEC     = "#64748b"   # Slate 500
SUCCESS      = "#059669"   # Emerald 600
ERROR_COL    = "#dc2626"   # Red 600
WARN         = "#d97706"   # Amber 600
BTN_LIGHT    = "#f8fafc"   # Secondary button bg
BTN_HOV      = "#e2e8f0"   # Secondary hover
WARN_BG      = "#fffbeb"   # Amber 50
WARN_HOV     = "#fef3c7"   # Amber 100

def create_rounded_rect(canvas, x1, y1, x2, y2, r, **kwargs):
    return canvas.create_polygon(
        x1+r, y1, x2-r, y1, x2, y1, x2, y1+r, x2, y2-r, x2, y2, x2-r, y2, 
        x1+r, y2, x1, y2, x1, y2-r, x1, y1+r, x1, y1,
        smooth=True, **kwargs
    )

class ModernRoundedButton(tk.Canvas):
    def __init__(self, parent, text, command=None, height=45, radius=8, 
                 bg_color=PRIMARY, hover_color=PRIMARY_HOV, text_color="#ffffff", icon=None):
        super().__init__(parent, height=height, bg=parent["bg"], highlightthickness=0)
        self.command = command
        self.bg_color = bg_color
        self.hover_color = hover_color
        self.radius = radius
        self.state = "normal"
        self.text_val = text
        self.text_color = text_color
        self.icon = icon
        self.attention_ring = False

        self.bind("<Configure>", self._on_resize)
        self.bind("<Button-1>", self._on_click)
        self.bind("<ButtonRelease-1>", self._on_release)
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)

    def set_attention_ring(self, on: bool):
        self.attention_ring = bool(on)
        base = self.bg_color if self.state == "normal" else BORDER
        self._draw(base)

    def _on_resize(self, event):
        self._draw(self.bg_color if self.state == "normal" else BORDER)
        
    def _draw(self, color):
        self.delete("all")
        w, h = self.winfo_width(), self.winfo_height()
        if w < 10 or h < 10: return
        if self.attention_ring:
            create_rounded_rect(
                self, 1, 1, w - 1, h - 1, self.radius, fill=color, outline=WARN, width=3,
            )
        else:
            create_rounded_rect(self, 1, 1, w - 1, h - 1, self.radius, fill=color, outline="")
        
        display_text = f"{self.icon}   {self.text_val}" if self.icon else self.text_val
        self.text_id = self.create_text(w/2, h/2, text=display_text, fill=self.text_color if self.state == "normal" else TEXT_SEC, font=("Segoe UI", 11, "bold"), justify="center")

    def _on_enter(self, e):
        if self.state == "normal": self._draw(self.hover_color)
    def _on_leave(self, e):
        if self.state == "normal": self._draw(self.bg_color)
    def _on_click(self, e): pass
    def _on_release(self, e):
        if self.state == "normal" and self.command:
            self._draw(self.bg_color)
            self.command()

    def configure(self, **kwargs):
        redraw = False
        if "attention_ring" in kwargs:
            self.attention_ring = bool(kwargs.pop("attention_ring"))
            redraw = True
        if "state" in kwargs:
            self.state = kwargs.pop("state")
            redraw = True
        if redraw:
            self._draw(self.bg_color if self.state == "normal" else BORDER)

class DbCard(tk.Frame):
    """Database 상태 + 부품(카테고리)별 건수 요약."""
    def __init__(self, parent):
        super().__init__(parent, bg=parent["bg"])
        self.card = tk.Frame(self, bg=BG_CARD, highlightthickness=1, highlightbackground=BORDER)
        self.card.pack(fill="x")

        top = tk.Frame(self.card, bg=BG_CARD)
        top.pack(fill="x", padx=14, pady=(12, 4))
        tk.Label(top, text="🗄", bg=BG_CARD, font=("Segoe UI", 14)).pack(side="left")
        tk.Label(top, text="Database", bg=BG_CARD, fg=TEXT_PRI, font=("Segoe UI", 12, "bold")).pack(side="left", padx=(8, 0))
        self.lbl_status = tk.Label(top, text="⏳ 대기 중…", bg=BG_CARD, fg=TEXT_SEC, font=("Segoe UI", 10, "bold"))
        self.lbl_status.pack(side="right")

        self.lbl_desc = tk.Label(
            self.card, text="데이터 대기 중", bg=BG_CARD, fg=TEXT_SEC,
            font=("Segoe UI", 10), justify="left", anchor="w", wraplength=520
        )
        self.lbl_desc.pack(fill="x", padx=14, pady=(0, 6))

        self.lbl_breakdown = tk.Label(
            self.card, text="", bg=BG_CARD, fg=TEXT_PRI,
            font=("Segoe UI", 10), justify="left", anchor="w", wraplength=520
        )
        self.lbl_breakdown.pack(fill="x", padx=14, pady=(0, 12))

    def set_loaded(self, records):
        total = len(records)
        br = category_breakdown_text(records)
        self.lbl_status.configure(text="✅ 로드 완료", fg=PRIMARY)
        self.lbl_desc.configure(text=f"총 {total}건이 로드되었습니다.", fg=TEXT_SEC)
        self.lbl_breakdown.configure(text=f"부품별: {br}", fg=TEXT_PRI)
        
    def set_error(self, err):
        self.lbl_status.configure(text="❌ 로드 실패", fg=ERROR_COL)
        self.lbl_desc.configure(text=str(err), fg=ERROR_COL)
        self.lbl_breakdown.configure(text="")

class DashedUploadDropZone(tk.Canvas):
    def __init__(self, parent, title_text="엑셀 파일을 드래그하여 업로드", icon_text="📊"):
        super().__init__(parent, bg=parent["bg"], highlightthickness=0)
        self.bind("<Configure>", self._on_resize)
        self.state_text = title_text
        self.sub_text = "또는 아래 영역을 클릭하세요"
        self.icon_text = icon_text
        self.is_selected = False
        self.is_loading = False
        self.fname = ""
        self.loading_pct = 0
        
    def _on_resize(self, e): self._draw()

    def _draw(self):
        self.delete("all")
        w, h = self.winfo_width(), self.winfo_height()
        if w < 10: return
        if self.is_loading:
            create_rounded_rect(self, 2, 2, w-2, h-2, 10, fill=PRIMARY_SOFT, outline=PRIMARY, width=2)
            self.create_text(w/2, h/2 - 26, text="⏳ 파일 업로드/파싱 중", fill=PRIMARY, font=("Segoe UI", 13, "bold"))
            self.create_text(w/2, h/2 + 2, text=f"{self.loading_pct}%", fill=PRIMARY, font=("Segoe UI", 12, "bold"))
            self.create_text(w/2, h/2 + 30, text=self.fname, fill=TEXT_PRI, font=("Segoe UI", 10))
        elif self.is_selected:
            create_rounded_rect(self, 2, 2, w-2, h-2, 10, fill=PRIMARY_SOFT, outline=PRIMARY, width=2)
            self.create_text(w/2, h/2 - 18, text="✓ 파일 준비 완료", fill=PRIMARY, font=("Segoe UI", 15, "bold"))
            self.create_text(w/2, h/2 + 22, text=self.fname, fill=TEXT_PRI, font=("Segoe UI", 11))
        else:
            create_rounded_rect(self, 2, 2, w-2, h-2, 10, fill=BG_CARD, outline=BORDER, dash=(8, 4), width=1.5)
            self.create_oval(w/2-28, h/2-44, w/2+28, h/2+12, fill=BG_MAIN, outline=BORDER, width=1)
            self.create_text(w/2, h/2-16, text=self.icon_text, font=("Segoe UI", 26))
            self.create_text(w/2, h/2 + 32, text=self.state_text, fill=TEXT_PRI, font=("Segoe UI", 13, "bold"))
            self.create_text(w/2, h/2 + 58, text=self.sub_text, fill=TEXT_SEC, font=("Segoe UI", 10))

    def set_selected(self, fname):
        self.is_loading = False
        self.is_selected = True
        self.fname = fname
        self._draw()

    def set_loading(self, fname, pct=0):
        self.is_selected = False
        self.is_loading = True
        self.fname = fname
        self.loading_pct = max(0, min(100, int(pct)))
        self._draw()

class LogCard(tk.Canvas):
    def __init__(self, parent, raw_msg, color_type):
        super().__init__(parent, height=65, bg=BG_MAIN, highlightthickness=0)
        self.raw_msg = raw_msg
        self.tag = "INF"
        self.content = raw_msg
        
        m = re.match(r'^\[(.*?)\] (.*)', raw_msg)
        if m:
            self.tag = m.group(1).upper()
            self.content = m.group(2)

        self.accent_color = BORDER
        self.label_color = TEXT_SEC
        
        if color_type == "red" or "오류" in self.tag or "실패" in self.tag:
            self.accent_color = ERROR_COL
            self.label_color = ERROR_COL
        elif color_type == "yellow" or "미일치" in self.tag:
            self.accent_color = WARN
            self.label_color = WARN
        elif "CAP" in self.tag or "IC" in self.tag or "MOSFET" in self.tag or "DIODE" in self.tag or "RECT" in self.tag or "일치" in self.tag or color_type == "green":
            self.accent_color = PRIMARY
            self.label_color = PRIMARY

        self.bind("<Configure>", self._on_resize)
        
    def _on_resize(self, e):
        self.delete("all")
        w, h = self.winfo_width(), self.winfo_height()
        if w < 10: return
        create_rounded_rect(self, 2, 2, w-2, h-2, 8, fill=BG_CARD, outline=BORDER, width=1)
        self.create_rectangle(2, 12, 5, h-12, fill=self.accent_color, outline="")
        self.create_text(18, 20, text=self.tag, fill=self.label_color, font=("Segoe UI", 9, "bold"), anchor="w")
        self.create_text(18, 42, text=self.content[:100] + ('…' if len(self.content) > 100 else ''), fill=TEXT_PRI, font=("Consolas", 10), anchor="w")

class ScrollableLogFrame(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=BG_MAIN)
        self.canvas = tk.Canvas(self, bg=BG_MAIN, highlightthickness=0)
        style = ttk.Style()
        # Modern Light Scrollbar
        style.configure("Light.Vertical.TScrollbar", background=BG_CARD, troughcolor=BG_MAIN, bordercolor=BORDER, arrowcolor=PRIMARY)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview, style="Light.Vertical.TScrollbar")
        self.scrollable_frame = tk.Frame(self.canvas, bg=BG_MAIN)
        self.window_id = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y", padx=(5,0))
        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.bind("<Configure>", lambda e: self.canvas.itemconfig(self.window_id, width=e.width))
        self.log_cards = []

    def add_log(self, text, color_type=None):
        card = LogCard(self.scrollable_frame, text, color_type)
        card.pack(fill="x", padx=2, pady=(0, 8))
        self.log_cards.append(card)
        if len(self.log_cards) > 500:
            old = self.log_cards.pop(0)
            old.destroy()
        self.canvas.update_idletasks()
        self.canvas.yview_moveto(1)

        return len(self.log_cards)

    def clear_all(self):
        for w in self.scrollable_frame.winfo_children():
            w.destroy()
        self.log_cards = []
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

class FlatStatusBar(tk.Canvas):
    def __init__(self, parent):
        super().__init__(parent, height=40, bg=BG_CARD, highlightthickness=0)
        self.bind("<Configure>", self._on_resize)
        self.indicator_color = TEXT_SEC
        self.status_text = "준비"
        self.pos = 0
        self.is_running = False
        self.bar_width = 150
        
    def _on_resize(self, e): self._draw()
        
    def _draw(self):
        self.delete("all")
        w, h = self.winfo_width(), self.winfo_height()
        if w < 10: return
        self.create_line(0, 0, w, 0, fill=BORDER, width=1)
        self.create_oval(24, h/2-4, 32, h/2+4, fill=self.indicator_color, outline="")
        self.create_text(44, h/2, text=self.status_text, fill=TEXT_PRI, font=("Segoe UI", 10), anchor="w")
        self.create_text(w-24, h/2, text="Dongyang ENP Auto Spec v2.1", fill=TEXT_SEC, font=("Segoe UI", 9), anchor="e")
        
        if self.is_running:
            track_w = 200
            track_x = w/2 - track_w/2
            create_rounded_rect(self, track_x, h/2-4, track_x+track_w, h/2+4, 4, fill=BG_MAIN, outline=BORDER)
            bar_x = track_x + self.pos
            create_rounded_rect(self, bar_x, h/2-4, min(bar_x+self.bar_width, track_x+track_w), h/2+4, 4, fill=PRIMARY, outline="")

    def set_status(self, msg, color=TEXT_SEC):
        self.status_text = msg
        self.indicator_color = color
        self._draw()

    def start_progress(self):
        self.is_running = True
        self.pos = 0
        self.set_status("처리 중...", WARN)
        self._animate()

    def _animate(self):
        if not self.is_running: return
        self.pos += 4
        if self.pos > 200: self.pos = -self.bar_width
        self._draw()
        self.after(20, self._animate)
        
    def stop_progress(self, msg, success=True):
        self.is_running = False
        self.set_status(msg, SUCCESS if success else ERROR_COL)

class AddPartDialog(tk.Toplevel):
    """신규 부품 등록 → DB/Database.xlsx 반영."""
    def __init__(self, parent, on_success=None):
        super().__init__(parent)
        self.on_success = on_success
        self.title("신규 부품 등록")
        self.configure(bg=BG_MAIN)
        self.geometry("460x400")
        self.minsize(420, 360)
        self.transient(parent)
        self.grab_set()

        frm = tk.Frame(self, bg=BG_MAIN, padx=18, pady=16)
        frm.pack(fill="both", expand=True)
        frm.grid_columnconfigure(0, weight=1)

        tk.Label(frm, text="카테고리", bg=BG_MAIN, fg=TEXT_PRI, font=("Segoe UI", 10)).grid(row=0, column=0, sticky="w", pady=(0, 4))
        self.cb_cat = ttk.Combobox(frm, values=list(DB_CATEGORY_SCHEMA.keys()), state="readonly", width=20)
        self.cb_cat.set("IC")
        self.cb_cat.grid(row=1, column=0, sticky="ew", pady=(0, 12))
        self.cb_cat.bind("<<ComboboxSelected>>", self._rebuild_specs)

        tk.Label(frm, text="부품명 (시트 C열)", bg=BG_MAIN, fg=TEXT_PRI, font=("Segoe UI", 10)).grid(row=2, column=0, sticky="w", pady=(0, 4))
        self.ent_pn = tk.Entry(frm, font=("Segoe UI", 10))
        self.ent_pn.grid(row=3, column=0, sticky="ew", pady=(0, 12))

        self.spec_frame = tk.Frame(frm, bg=BG_MAIN)
        self.spec_frame.grid(row=4, column=0, sticky="ew", pady=(0, 12))
        self.spec_entries = []

        btn_row = tk.Frame(frm, bg=BG_MAIN)
        btn_row.grid(row=5, column=0, sticky="ew", pady=(8, 0))
        ModernRoundedButton(btn_row, "등록", self._submit, height=38, radius=8, bg_color=PRIMARY, hover_color=PRIMARY_HOV, text_color="#ffffff", icon="✓").pack(side="left", padx=(0, 8))
        ModernRoundedButton(btn_row, "취소", lambda: self.destroy(), height=38, radius=8, bg_color=BTN_LIGHT, hover_color=BTN_HOV, text_color=TEXT_PRI).pack(side="left")

        self._rebuild_specs()

    def _rebuild_specs(self, event=None):
        for w in self.spec_frame.winfo_children():
            w.destroy()
        self.spec_entries.clear()
        cat = self.cb_cat.get()
        labels = get_category_spec_labels(cat)
        for label in labels:
            tk.Label(self.spec_frame, text=label, bg=BG_MAIN, fg=TEXT_SEC, font=("Segoe UI", 9)).pack(anchor="w")
            e = tk.Entry(self.spec_frame, font=("Segoe UI", 10))
            e.pack(fill="x", pady=(0, 8))
            self.spec_entries.append(e)

    def _submit(self):
        cat = self.cb_cat.get()
        part_name = self.ent_pn.get().strip()
        specs = [e.get().strip() for e in self.spec_entries]
        if not any(specs):
            messagebox.showwarning("입력 확인", "스펙을 최소 한 칸 이상 입력하세요.")
            return
        try:
            append_part_to_database(cat, part_name, specs)
            messagebox.showinfo("완료", f"Database.xlsx에 등록되었습니다.\n{DB_FILE}")
            if self.on_success:
                self.on_success()
            self.destroy()
        except Exception as ex:
            messagebox.showerror("등록 실패", str(ex))

# ─────────────────────────────────────────────
# 메인 어플리케이션
# ─────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Auto Spec Program")
        self.state('zoomed')
        self.minsize(1024, 768)
        self.configure(bg=BG_MAIN)

        self.app_config = load_app_config()
        self.log_queue = queue.Queue()
        self._log_history = []
        self.db_records = []
        self.pl_lookup = {}
        self.pl_paths = []
        self.pl_vendor_rank_by_loc = {}
        self.form_locations = set()
        self.pl_loading = False
        self.run_loading = False
        self.log_win = None
        self.log_container = None
        self.unseen_log_count = 0
        self.excel_path = None
        self.last_unmatched = []
        self.batch_paths = []
        self.pl_vendor_confirmed = False
        self._vendor_pulse_job = None
        # Part List: 같은 실행 중 동일 파일 조합 재파싱 방지(결과 동일, 속도만 개선)
        # key = (abs_path, size, mtime) 튜플들의 튜플
        self._pl_mem_cache = {}
        self._pl_last_sig = None

        self._build_ui()
        self._sync_vendor_button_state()
        self._setup_dnd()
        self._bind_shortcuts()
        self._process_log_queue()
        self._load_db_on_start()
        self._refresh_recent_combo()

    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=0)
        self.grid_rowconfigure(2, weight=1)

        header = tk.Frame(self, bg=BG_HEADER)
        header.grid(row=0, column=0, columnspan=2, sticky="ew", padx=0, pady=0)
        header.grid_columnconfigure(0, weight=1)
        
        header_inner = tk.Frame(header, bg=BG_HEADER)
        header_inner.pack(fill="x", padx=40, pady=(20, 16))
        
        logo_loaded = False
        if HAS_PIL:
            logo_path = os.path.join(BASE_DIR, "image", "logo.ico")
            if os.path.exists(logo_path):
                try:
                    img = Image.open(logo_path)
                    w, h = img.size
                    new_h = 40
                    new_w = int(w * (new_h / h))
                    try:
                        resample = Image.Resampling.LANCZOS
                    except AttributeError:
                        resample = Image.LANCZOS
                    img = img.resize((new_w, new_h), resample)
                    self.logo_img = ImageTk.PhotoImage(img)
                    logo_lbl = tk.Label(header_inner, image=self.logo_img, bg=BG_HEADER, bd=0)
                    logo_lbl.pack(side="left", padx=(0, 14))
                    logo_loaded = True
                except Exception:
                    pass

        if not logo_loaded:
            logo_frame = tk.Frame(header_inner, bg=PRIMARY_SOFT, width=40, height=40)
            logo_frame.pack(side="left", padx=(0, 14))
            logo_frame.pack_propagate(False)
            logo_lbl = tk.Label(logo_frame, text="DY", fg=PRIMARY, bg=PRIMARY_SOFT, font=("Segoe UI", 14, "bold"))
            logo_lbl.place(relx=0.5, rely=0.5, anchor="center")
            
        title_box = tk.Frame(header_inner, bg=BG_HEADER)
        title_box.pack(side="left")
        tk.Label(title_box, text="Dongyang ENP Auto Spec", fg=TEXT_PRI, bg=BG_HEADER, font=("Segoe UI", 17, "bold")).pack(anchor="w")
        tk.Label(title_box, text="데이터베이스 기준으로 엑셀 측정값을 자동 검사하고 보고서를 생성합니다", fg=TEXT_SEC, bg=BG_HEADER, font=("Segoe UI", 10)).pack(anchor="w", pady=(2, 0))
        
        ver_frame = tk.Frame(header_inner, bg=BORDER, padx=1, pady=1)
        ver_frame.pack(side="right", fill="y", pady=2)
        tk.Label(ver_frame, text="v2.1", fg=TEXT_SEC, bg=BG_CARD, font=("Segoe UI", 9, "bold"), padx=10, pady=3).pack()

        sep = tk.Frame(self, height=1, bg=BORDER)
        sep.grid(row=1, column=0, columnspan=2, sticky="ew")

        content_wrap = tk.Frame(self, bg=BG_MAIN)
        content_wrap.grid(row=2, column=0, columnspan=2, sticky="nsew")
        content_wrap.grid_columnconfigure(0, weight=1)
        content_wrap.grid_rowconfigure(0, weight=1)

        # 메인 컨텐츠를 가운데 정렬한 단일 컬럼 레이아웃
        left_panel = tk.Frame(content_wrap, bg=BG_MAIN)
        left_panel.grid(row=0, column=0, sticky="n", padx=(40, 40), pady=(0, 20))
        left_panel.grid_rowconfigure(1, weight=1)

        db_container = tk.Frame(left_panel, bg=BG_MAIN)
        db_container.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        self.db_card = DbCard(db_container)
        self.db_card.pack(fill="x")
        self.btn_add_part = ModernRoundedButton(
            db_container, "부품 등록", self._show_add_part_dialog,
            height=38, radius=10, bg_color=PRIMARY_SOFT, hover_color="#dbeafe", text_color=PRIMARY, icon="＋"
        )
        self.btn_add_part.pack(fill="x", pady=(10, 0))
        self.btn_add_part.configure(state="disabled")

        pl_row = tk.Frame(db_container, bg=BG_MAIN)
        pl_row.pack(fill="x", pady=(8, 0))
        self.lbl_pl_status = tk.Label(
            pl_row, text="Part List: 대기 중…", bg=BG_MAIN, fg=TEXT_SEC,
            font=("Segoe UI", 9), anchor="w", justify="left", wraplength=480
        )
        self.lbl_pl_status.pack(side="left", fill="x", expand=True)
        self.pb_pl = ttk.Progressbar(pl_row, orient="horizontal", mode="determinate", length=120, maximum=100)
        self.pb_pl.pack(side="right", padx=(4, 4))
        self.lbl_pl_percent = tk.Label(
            pl_row, text="0%", bg=BG_MAIN, fg=TEXT_SEC, font=("Segoe UI", 9, "bold"), width=5, anchor="e"
        )
        self.lbl_pl_percent.pack(side="right", padx=(6, 4))
        pl_vendor_row = tk.Frame(db_container, bg=BG_MAIN)
        pl_vendor_row.pack(fill="x", pady=(10, 0))
        self.btn_vendor = ModernRoundedButton(
            pl_vendor_row,
            "벤더 선택 — 로케이션별 1·2·3차 (Part List 파싱 후 필수)",
            self._open_vendor_selector,
            height=48,
            radius=10,
            bg_color=PRIMARY,
            hover_color=PRIMARY_HOV,
            text_color="#ffffff",
            icon="▶",
        )
        self.btn_vendor.pack(fill="x")
        upload_wrap = tk.Frame(left_panel, bg=BG_MAIN)
        upload_wrap.grid(row=1, column=0, sticky="nsew", pady=(0, 12))
        upload_wrap.grid_columnconfigure(0, weight=1)
        upload_wrap.grid_columnconfigure(1, weight=1)
        upload_wrap.grid_rowconfigure(0, weight=1)

        self.upload_zone = DashedUploadDropZone(upload_wrap, "양식 파일 업로드", "📊")
        self.upload_zone.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        self.upload_zone.bind("<Button-1>", lambda e: self._select_file())

        self.pl_upload_zone = DashedUploadDropZone(upload_wrap, "Part List 파일 업로드", "🧩")
        self.pl_upload_zone.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        self.pl_upload_zone.bind("<Button-1>", lambda e: self._select_part_list_files())

        optf = tk.Frame(left_panel, bg=BG_MAIN)
        optf.grid(row=2, column=0, sticky="ew", pady=(0, 10))
        tk.Label(optf, text="전처리·옵션", bg=BG_MAIN, fg=TEXT_SEC, font=("Segoe UI", 9, "bold")).pack(anchor="w")
        self.var_open_result = tk.BooleanVar(value=self.app_config.get("open_result_after_run", True))
        # 전처리 옵션은 기본 탑재(항상 ON)로 두고 UI에는 완료 후 폴더 열기만 노출
        self.var_norm = tk.BooleanVar(value=True)
        self.var_validate = tk.BooleanVar(value=True)
        self.var_dry_run = tk.BooleanVar(value=False)
        self.var_pl_lookup = tk.BooleanVar(value=True)
        tk.Checkbutton(
            optf, text="처리 후 결과 폴더 자동 열기", variable=self.var_open_result,
            bg=BG_MAIN, fg=TEXT_PRI, selectcolor=BG_CARD,
            font=("Segoe UI", 9), activebackground=BG_MAIN, activeforeground=TEXT_PRI,
            command=self._persist_config,
        ).pack(anchor="w", pady=1)

        recent_fr = tk.Frame(left_panel, bg=BG_MAIN)
        recent_fr.grid(row=3, column=0, sticky="ew", pady=(0, 12))
        tk.Label(recent_fr, text="최근 파일", bg=BG_MAIN, fg=TEXT_SEC, font=("Segoe UI", 9)).pack(anchor="w")
        self.cmb_recent = ttk.Combobox(recent_fr, state="readonly", width=58)
        self.cmb_recent.pack(fill="x", pady=(4, 0))
        self.cmb_recent.bind("<<ComboboxSelected>>", self._on_recent_selected)

        btn_grid = tk.Frame(left_panel, bg=BG_MAIN)
        btn_grid.grid(row=4, column=0, sticky="ew")
        btn_grid.grid_columnconfigure(0, weight=1)
        btn_grid.grid_columnconfigure(1, weight=1)

        self.btn_run = ModernRoundedButton(btn_grid, "엑셀 자동 입력", self._run, height=44, radius=10, bg_color=PRIMARY, hover_color=PRIMARY_HOV, text_color="#ffffff", icon="▶")
        self.btn_run.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        self.btn_open = ModernRoundedButton(btn_grid, "결과 폴더", self._open_result_folder, height=44, radius=10, bg_color=PRIMARY_SOFT, hover_color="#dbeafe", text_color=PRIMARY, icon="📁")
        self.btn_open.grid(row=0, column=1, sticky="ew")
        self.btn_log_view = ModernRoundedButton(
            btn_grid, "로그 보기 (0)", self._open_log_window,
            height=36, radius=10, bg_color=PRIMARY_SOFT, hover_color="#dbeafe", text_color=PRIMARY, icon="🧾"
        )
        self.btn_log_view.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(8, 0))

        run_prog_row = tk.Frame(left_panel, bg=BG_MAIN)
        run_prog_row.grid(row=5, column=0, sticky="ew", pady=(10, 0))
        tk.Label(run_prog_row, text="최종 처리 진행률", bg=BG_MAIN, fg=TEXT_SEC, font=("Segoe UI", 9)).pack(side="left")
        self.lbl_run_percent = tk.Label(
            run_prog_row, text="0%", bg=BG_MAIN, fg=TEXT_SEC, font=("Segoe UI", 9, "bold"), width=5, anchor="e"
        )
        self.lbl_run_percent.pack(side="right")
        self.pb_run = ttk.Progressbar(run_prog_row, orient="horizontal", mode="determinate", maximum=100)
        self.pb_run.pack(side="right", fill="x", expand=True, padx=(8, 8))

        self.btn_unmatched = ModernRoundedButton(left_panel, "미일치 항목 분석", self._show_unmatched, height=40, radius=10, bg_color=WARN_BG, hover_color=WARN_HOV, text_color=WARN, icon="⚠")
        self.btn_unmatched.grid(row=6, column=0, sticky="ew", pady=(12, 0))
        self.btn_unmatched.configure(state="disabled")

        self.status_bar = FlatStatusBar(self)
        self.status_bar.grid(row=3, column=0, columnspan=2, sticky="ew")

    def _setup_dnd(self):
        try:
            from tkinterdnd2 import DND_FILES
            self.upload_zone.drop_target_register(DND_FILES)
            self.upload_zone.dnd_bind("<<Drop>>", self._on_drop_main)
            self.pl_upload_zone.drop_target_register(DND_FILES)
            self.pl_upload_zone.dnd_bind("<<Drop>>", self._on_drop_pl)
        except Exception:
            pass

    def _bind_shortcuts(self):
        self.bind("<Control-o>", lambda e: self._select_file())
        self.bind("<Control-O>", lambda e: self._select_file())
        self.bind("<Control-Return>", lambda e: self._run())
        self.bind("<F5>", lambda e: self._run())
        self.bind("<Control-e>", lambda e: self._export_log())
        self.bind("<Control-E>", lambda e: self._export_log())
        self.bind("<Control-l>", lambda e: self._copy_log_to_clipboard())
        self.bind("<Control-L>", lambda e: self._copy_log_to_clipboard())

    def _persist_config(self):
        self.app_config["open_result_after_run"] = self.var_open_result.get()
        # 고정 기본값(UI 미노출)
        self.var_norm.set(True)
        self.var_validate.set(True)
        self.var_dry_run.set(False)
        self.var_pl_lookup.set(True)
        self.app_config["normalize_cell_text"] = True
        self.app_config["validate_workbook_on_run"] = True
        self.app_config["dry_run"] = False
        self.app_config["use_part_list_lookup"] = True
        save_app_config(self.app_config)

    def _refresh_recent_combo(self):
        paths = [p for p in self.app_config.get("recent_files", []) if os.path.isfile(p)]
        self.cmb_recent.unbind("<<ComboboxSelected>>")
        self.cmb_recent["values"] = paths
        if paths:
            self.cmb_recent.current(0)
        else:
            self.cmb_recent.set("")
        self.cmb_recent.bind("<<ComboboxSelected>>", self._on_recent_selected)

    def _on_recent_selected(self, event=None):
        p = self.cmb_recent.get()
        if p and os.path.isfile(p):
            self.batch_paths = []
            self._set_file(p, add_recent=False)

    def _select_batch_files(self):
        from tkinter import filedialog
        paths = filedialog.askopenfilenames(
            title="일괄 처리할 엑셀 선택",
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("모든 파일", "*.*")]
        )
        if not paths:
            return
        self.batch_paths = list(paths)
        self._set_file(self.batch_paths[0])
        self._log(f"[안내] 일괄 처리 대기: {len(self.batch_paths)}개 파일 (처리 시작 시 순차 실행)", "green")

    def _paste_path_from_clipboard(self):
        try:
            raw = self.clipboard_get()
        except Exception:
            messagebox.showinfo("안내", "클립보드가 비어 있거나 읽을 수 없습니다.")
            return
        p = raw.strip().strip('"').strip("'")
        if os.path.isfile(p) and p.lower().endswith((".xlsx", ".xlsm")):
            self.batch_paths = []
            self._set_file(p)
        else:
            messagebox.showwarning("안내", "클립보드에 유효한 .xlsx / .xlsm 경로가 없습니다.")

    def _export_log(self):
        from tkinter import filedialog
        if not self._log_history:
            messagebox.showinfo("안내", "저장할 로그가 없습니다.")
            return
        path = filedialog.asksaveasfilename(
            title="로그 저장",
            defaultextension=".txt",
            filetypes=[("텍스트", "*.txt"), ("모든 파일", "*.*")]
        )
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                for msg, _ in self._log_history:
                    f.write(msg + "\n")
            self._log(f"[안내] 로그 저장: {path}", "green")
        except Exception as e:
            messagebox.showerror("오류", str(e))

    def _clear_log(self):
        self._log_history.clear()
        self.unseen_log_count = 0
        if self.log_container:
            self.log_container.clear_all()
        if hasattr(self, "btn_log_view"):
            self.btn_log_view.text_val = "로그 보기 (0)"
            self.btn_log_view._draw(self.btn_log_view.bg_color)

    def _copy_log_to_clipboard(self):
        if not self._log_history:
            messagebox.showinfo("안내", "복사할 로그가 없습니다.")
            return
        text = "\n".join(m for m, _ in self._log_history)
        self.clipboard_clear()
        self.clipboard_append(text)
        self._log("[안내] 로그를 클립보드에 복사했습니다.", "green")

    def _open_log_window(self):
        if self.log_win and self.log_win.winfo_exists():
            self.log_win.lift()
            return
        self.log_win = tk.Toplevel(self)
        self.log_win.title("터미널 로그")
        self.log_win.geometry("900x620")
        self.log_win.configure(bg=BG_MAIN)

        hdr = tk.Frame(self.log_win, bg=BG_MAIN)
        hdr.pack(fill="x", padx=14, pady=(12, 8))
        tk.Label(hdr, text="터미널 로그", fg=TEXT_PRI, bg=BG_MAIN, font=("Segoe UI", 12, "bold")).pack(side="left")
        for lbl, cmd in (("저장", self._export_log), ("지우기", self._clear_log), ("복사", self._copy_log_to_clipboard)):
            tk.Button(
                hdr, text=lbl, command=cmd, relief=tk.FLAT, bg=BG_MAIN, fg=PRIMARY,
                cursor="hand2", font=("Segoe UI", 9, "underline"), padx=6, bd=0
            ).pack(side="right")

        self.log_container = ScrollableLogFrame(self.log_win)
        self.log_container.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        for msg, color in self._log_history:
            self.log_container.add_log(msg, color)

        self.unseen_log_count = 0
        self.btn_log_view.text_val = "로그 보기 (0)"
        self.btn_log_view._draw(self.btn_log_view.bg_color)

        def _on_close():
            self.log_container = None
            self.log_win.destroy()
            self.log_win = None
        self.log_win.protocol("WM_DELETE_WINDOW", _on_close)

    def _parse_drop_paths(self, event_data):
        try:
            parts = list(self.tk.splitlist(event_data))
        except Exception:
            parts = [event_data]
        out = []
        for p in parts:
            p = str(p).strip().strip("{}").strip('"').strip("'")
            if p:
                out.append(p)
        return out

    def _on_drop_main(self, event):
        paths = self._parse_drop_paths(event.data)
        if not paths:
            return
        self.batch_paths = []
        self._set_file(paths[0])

    def _on_drop_pl(self, event):
        paths = self._parse_drop_paths(event.data)
        if not paths:
            return
        self._set_part_list_files(paths)

    def _select_file(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx *.xlsm"), ("모든 파일", "*.*")]
        )
        if path:
            self.batch_paths = []
            self._set_file(path)

    def _set_file(self, path, add_recent=True):
        self.excel_path = path
        fname = os.path.basename(path)
        self.upload_zone.set_selected(fname)
        self.form_locations = self._extract_form_locations(path)
        if self.pl_lookup:
            self.pl_vendor_confirmed = False
        if add_recent:
            self.app_config = add_recent_file(self.app_config, path)
            self._refresh_recent_combo()
        self._log(f"[안내] 엑셀 파일 선택 완료: {fname} (로케이션 {len(self.form_locations)}개 탐지)", "green")
        self._sync_vendor_button_state()

    def _extract_form_locations(self, path):
        out = set()
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            max_r = min(ws.max_row or 14, 10000)
            for row_cells in ws.iter_rows(min_row=14, max_row=max_r, min_col=13, max_col=18, values_only=True):
                for cell in row_cells:
                    if not cell:
                        continue
                    txt = str(cell).replace("\n", " ").replace("\r", " ")
                    for tok in re.split(r"[\s,;/]+", txt):
                        tok = tok.strip().upper()
                        if len(tok) >= 3:
                            out.add(tok)
            wb.close()
        except Exception as e:
            self._log(f"[안내] 양식 로케이션 추출 실패: {e}", "yellow")
        return out

    def _select_part_list_files(self):
        from tkinter import filedialog
        paths = filedialog.askopenfilenames(
            title="Part List 파일 선택 (복수 가능)",
            filetypes=[("Excel 파일", "*.xlsx *.xlsm"), ("모든 파일", "*.*")]
        )
        if not paths:
            return
        self._set_part_list_files(paths)

    def _pl_signature(self, paths):
        sig = []
        for p in paths or []:
            try:
                ap = os.path.abspath(os.path.normpath(p))
                st = os.stat(ap)
                sig.append((ap, int(st.st_size), int(st.st_mtime)))
            except Exception:
                ap = os.path.abspath(os.path.normpath(p))
                sig.append((ap, 0, 0))
        sig.sort(key=lambda t: t[0].lower())
        return tuple(sig)

    def _set_part_list_files(self, paths):
        cleaned = []
        for p in paths:
            if os.path.isfile(p) and p.lower().endswith((".xlsx", ".xlsm")):
                cleaned.append(os.path.normpath(p))
        if not cleaned:
            messagebox.showwarning("안내", "유효한 Part List 엑셀 파일(.xlsx/.xlsm)을 찾지 못했습니다.")
            return
        self.pl_vendor_confirmed = False
        # 동일 파일 조합이면, 이미 로드된 결과/캐시를 재사용 (결과 동일)
        sig = self._pl_signature(cleaned)
        if (
            sig
            and sig == getattr(self, "_pl_last_sig", None)
            and self.pl_lookup
            and not self.pl_loading
        ):
            self._log("[PL] 동일 Part List 조합 — 재파싱 생략", "green")
            self._sync_vendor_button_state()
            return
        if sig and sig in self._pl_mem_cache and not self.pl_loading:
            cached = self._pl_mem_cache[sig]
            self.pl_lookup = cached.get("pl_lookup") or {}
            self.pl_vendor_rank_by_loc = cached.get("pl_vendor_rank_by_loc") or {}
            self.pl_paths = list(cleaned)
            self._pl_last_sig = sig
            nloc = len(self.pl_lookup)
            nent = sum(len(v) for v in self.pl_lookup.values())
            label = os.path.basename(cleaned[0]) if len(cleaned) == 1 else f"{len(cleaned)}개 파일"
            self.pl_upload_zone.set_selected(label)
            self.lbl_pl_status.configure(
                text=f"Part List: 로드 완료(캐시) - 로케이션 {nloc}개 / 후보 {nent}건",
                fg=PRIMARY,
            )
            self._set_pl_progress(100)
            self._log("[PL] 메모리 캐시 사용 — 즉시 로드", "green")
            self._sync_vendor_button_state()
            return

        self.pl_paths = cleaned
        self._pl_last_sig = sig
        label = os.path.basename(cleaned[0]) if len(cleaned) == 1 else f"{len(cleaned)}개 파일"
        self.pl_upload_zone.set_loading(label, 0)
        self._log(f"[PL] 선택 파일 {len(self.pl_paths)}개 지정", "green")
        self.pl_loading = True
        self._set_pl_progress(0)
        self.lbl_pl_status.configure(text="Part List: 파일 업로드 중... 0%", fg=TEXT_SEC)
        threading.Thread(target=self._load_part_list_worker, daemon=True).start()

    def _open_vendor_selector(self):
        if not self.pl_lookup:
            messagebox.showinfo("안내", "먼저 Part List 파일을 업로드/로드한 뒤 벤더를 선택하세요.")
            return
        if not self.form_locations:
            messagebox.showinfo("안내", "먼저 양식 파일을 업로드하세요. (M~R 로케이션 추출 필요)")
            return
        win = tk.Toplevel(self)
        win.title("벤더 선택 (1차/2차/3차)")
        win.geometry("900x620")
        win.configure(bg=BG_MAIN)

        top = tk.Frame(win, bg=BG_MAIN)
        top.pack(fill="x", padx=12, pady=(10, 6))
        tk.Label(top, text="로케이션별 벤더 선택", bg=BG_MAIN, fg=TEXT_PRI, font=("Segoe UI", 11, "bold")).pack(side="left")
        tk.Label(top, text="(자동=미지정, 후보 순서 그대로)", bg=BG_MAIN, fg=TEXT_SEC, font=("Segoe UI", 9)).pack(side="left", padx=(8, 0))
        search_var = tk.StringVar(value="")
        tk.Label(top, text="검색", bg=BG_MAIN, fg=TEXT_SEC, font=("Segoe UI", 9)).pack(side="right")
        ent_search = tk.Entry(top, textvariable=search_var, font=("Consolas", 10), width=18)
        ent_search.pack(side="right", padx=(6, 10))

        body = tk.Frame(win, bg=BG_MAIN)
        body.pack(fill="both", expand=True, padx=12, pady=(0, 8))
        body.grid_columnconfigure(0, weight=3)
        body.grid_columnconfigure(1, weight=4)
        body.grid_rowconfigure(0, weight=1)

        left = tk.Frame(body, bg=BG_CARD, highlightthickness=1, highlightbackground=BORDER)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        right = tk.Frame(body, bg=BG_CARD, highlightthickness=1, highlightbackground=BORDER)
        right.grid(row=0, column=1, sticky="nsew")

        lb = tk.Listbox(left, font=("Consolas", 10))
        sb = ttk.Scrollbar(left, orient="vertical", command=lb.yview)
        lb.configure(yscrollcommand=sb.set)
        lb.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=8)
        sb.pack(side="right", fill="y", pady=8, padx=(0, 8))

        def _loc_sort_key(k):
            # 후보 1개(1차 고정)는 사용자 선택이 필요 없으므로 목록 맨 아래로 둔다.
            cands = self.pl_lookup.get(k, [])
            fixed_single = len(cands) <= 1
            return (1 if fixed_single else 0, k)

        all_keys = sorted(
            [k for k in self.pl_lookup.keys() if k in self.form_locations],
            key=_loc_sort_key,
        )
        filtered_keys = list(all_keys)
        idx_to_key = {}

        def _split_vendor_spec(raw):
            s = str(raw or "").strip()
            if not s:
                return "", ""
            if "||" in s:
                v, n = s.split("||", 1)
                return v.strip(), n.strip()
            # 세미콜론/기호가 앞에 붙은 케이스 정리 후 가능한 전체 품목명을 보여준다.
            s = s.lstrip(":;, ")
            if ";" in s:
                parts = [p.strip() for p in s.split(";") if p.strip()]
                s = parts[-1] if parts else s
            s = s.strip()
            if not s:
                s = str(raw or "").strip()
            toks = s.split()
            if not toks:
                return "", ""
            vendor = toks[0]
            body = " ".join(toks[1:]).strip()
            return vendor, body

        info_head = tk.Label(
            right, text="로케이션을 선택하세요", bg=BG_CARD, fg=TEXT_PRI,
            justify="left", anchor="w", font=("Consolas", 10)
        )
        info_head.pack(fill="x", padx=10, pady=(10, 6))
        cand_wrap = tk.Frame(right, bg=BG_CARD)
        cand_wrap.pack(fill="both", expand=True, padx=10, pady=(0, 6))
        cand_canvas = tk.Canvas(cand_wrap, bg=BG_CARD, highlightthickness=0)
        cand_sb = ttk.Scrollbar(cand_wrap, orient="vertical", command=cand_canvas.yview)
        cand_inner = tk.Frame(cand_canvas, bg=BG_CARD)
        cand_win = cand_canvas.create_window((0, 0), window=cand_inner, anchor="nw")
        cand_canvas.configure(yscrollcommand=cand_sb.set)
        cand_canvas.pack(side="left", fill="both", expand=True)
        cand_sb.pack(side="right", fill="y")
        cand_inner.bind("<Configure>", lambda e: cand_canvas.configure(scrollregion=cand_canvas.bbox("all")))
        cand_canvas.bind("<Configure>", lambda e: cand_canvas.itemconfig(cand_win, width=e.width))
        ctrl = tk.Frame(right, bg=BG_CARD)
        ctrl.pack(fill="x", padx=10, pady=(0, 10))
        tk.Label(ctrl, text="오른쪽 후보를 클릭하면 즉시 적용됩니다", bg=BG_CARD, fg=TEXT_SEC, font=("Segoe UI", 9)).pack(side="left")
        tk.Button(
            ctrl, text="초기화", relief=tk.FLAT, bg=PRIMARY_SOFT, fg=PRIMARY, padx=10,
            command=lambda: _set_auto_for_current()
        ).pack(side="right")

        def _confirm_vendor_and_close():
            self.pl_vendor_confirmed = True
            win.destroy()
            self._sync_vendor_button_state()

        tk.Button(
            ctrl, text="완료", relief=tk.FLAT, bg=PRIMARY, fg="#ffffff", padx=12,
            command=_confirm_vendor_and_close,
        ).pack(side="right", padx=(0, 8))
        tk.Button(
            ctrl, text="이전으로(Ctrl+Z)", relief=tk.FLAT, bg=BG_CARD, fg=PRIMARY, padx=8,
            command=lambda: _undo_last()
        ).pack(side="right", padx=(0, 8))

        def _render_list():
            lb.delete(0, "end")
            idx_to_key.clear()
            for i, k in enumerate(filtered_keys):
                cands = self.pl_lookup.get(k, [])
                if len(cands) == 1:
                    rank_txt = "1차(고정)"
                else:
                    rank = self.pl_vendor_rank_by_loc.get(k, 0)
                    rank_txt = "자동" if rank <= 0 else f"{rank}차"
                lb.insert("end", f"{k:16} | {rank_txt}")
                idx_to_key[i] = k

        def _refresh_row_by_key(key):
            for i, k in idx_to_key.items():
                if k != key:
                    continue
                cands = self.pl_lookup.get(k, [])
                if len(cands) == 1:
                    rank_txt = "1차(고정)"
                else:
                    rank = self.pl_vendor_rank_by_loc.get(k, 0)
                    rank_txt = "자동" if rank <= 0 else f"{rank}차"
                lb.delete(i)
                lb.insert(i, f"{k:16} | {rank_txt}")
                break

        current_key = {"value": None, "idx": -1}
        candidate_buttons = {}
        select_history = []

        def _find_next_index(start_idx):
            i = start_idx
            while i < lb.size():
                k = idx_to_key.get(i)
                if not k:
                    i += 1
                    continue
                cands = self.pl_lookup.get(k, [])
                # 1차 고정(후보 1개)은 자동 건너뜀
                if len(cands) <= 1:
                    i += 1
                    continue
                return i
            return -1

        def _select_index(i):
            if i < 0 or i >= lb.size():
                return
            lb.selection_clear(0, "end")
            lb.selection_set(i)
            lb.see(i)
            current_key["idx"] = i
            k = idx_to_key.get(i)
            if k:
                _render_candidates(k)

        def _render_candidates(k):
            current_key["value"] = k
            cands = self.pl_lookup.get(k, [])
            if len(cands) == 1:
                rank = 1
                chosen_rank = "1차(고정)"
            else:
                rank = self.pl_vendor_rank_by_loc.get(k, 0)
                chosen_rank = f"{rank}차" if rank > 0 else "자동"
            info_head.configure(text=f"[{k}] 현재 선택: {chosen_rank}")
            for w in cand_inner.winfo_children():
                w.destroy()
            candidate_buttons.clear()
            for i, v in enumerate(cands):
                vendor, body = _split_vendor_spec(v)
                if body:
                    line = f"{i+1}차: [{vendor}] {body}"
                else:
                    line = f"{i+1}차: [{vendor}]"
                is_selected = (rank == (i + 1))
                is_single = (len(cands) == 1)
                btn = tk.Button(
                    cand_inner,
                    text=("✓ " + line + ("  (고정)" if is_single else "")) if is_selected else line,
                    command=lambda r=i + 1: _pick_candidate_rank(r),
                    relief=tk.FLAT,
                    anchor="w",
                    justify="left",
                    bg=(PRIMARY_SOFT if is_selected else BG_CARD),
                    fg=(PRIMARY if is_selected else TEXT_PRI),
                    activebackground=PRIMARY_SOFT,
                    activeforeground=PRIMARY,
                    padx=10,
                    pady=7,
                    font=("Consolas", 10),
                )
                if is_single:
                    btn.configure(state="disabled", disabledforeground=PRIMARY if is_selected else TEXT_SEC)
                btn.pack(fill="x", pady=(0, 6))
                candidate_buttons[i + 1] = btn

        def _on_pick(_event=None):
            sel = lb.curselection()
            if not sel:
                return
            idx = sel[0]
            current_key["idx"] = idx
            k = idx_to_key.get(idx)
            if not k:
                return
            _render_candidates(k)

        def _set_auto_for_current():
            k = current_key.get("value")
            if not k:
                return
            if len(self.pl_lookup.get(k, [])) <= 1:
                return
            prev_rank = self.pl_vendor_rank_by_loc.get(k, 0)
            self.pl_vendor_rank_by_loc.pop(k, None)
            select_history.append((k, prev_rank, 0))
            _refresh_row_by_key(k)
            _render_candidates(k)

        def _pick_candidate_rank(rank):
            k = current_key.get("value")
            if not k:
                return
            cands = self.pl_lookup.get(k, [])
            if len(cands) <= 1:
                return
            if rank <= len(cands):
                prev_rank = self.pl_vendor_rank_by_loc.get(k, 0)
                self.pl_vendor_rank_by_loc[k] = rank
                select_history.append((k, prev_rank, rank))
                _refresh_row_by_key(k)
                _render_candidates(k)
                next_idx = _find_next_index(current_key.get("idx", -1) + 1)
                if next_idx >= 0:
                    _select_index(next_idx)

        def _undo_last():
            if not select_history:
                return
            k, prev_rank, _new_rank = select_history.pop()
            if prev_rank <= 0:
                self.pl_vendor_rank_by_loc.pop(k, None)
            else:
                self.pl_vendor_rank_by_loc[k] = prev_rank
            _refresh_row_by_key(k)
            # 복원된 로케이션으로 포커스 이동
            for i, kk in idx_to_key.items():
                if kk == k:
                    _select_index(i)
                    break

        def _on_search(*_):
            q = search_var.get().strip().upper()
            if not q:
                filtered_keys[:] = all_keys
            else:
                filtered_keys[:] = [k for k in all_keys if q in k]
            _render_list()
            if lb.size() > 0:
                _select_index(0)

        search_var.trace_add("write", _on_search)
        _render_list()
        if lb.size() > 0:
            first_idx = _find_next_index(0)
            _select_index(first_idx if first_idx >= 0 else 0)

        lb.bind("<<ListboxSelect>>", _on_pick)
        win.bind("<Control-z>", lambda e: _undo_last())
        win.bind("<Control-Z>", lambda e: _undo_last())

    def _set_pl_progress(self, pct):
        pct = max(0, min(100, int(pct)))
        self.pb_pl["value"] = pct
        self.lbl_pl_percent.configure(text=f"{pct}%")
        if self.pl_loading and pct < 100:
            self.lbl_pl_status.configure(text=f"Part List: 파일 업로드 중... {pct}%", fg=TEXT_SEC)
            if self.pl_paths:
                label = os.path.basename(self.pl_paths[0]) if len(self.pl_paths) == 1 else f"{len(self.pl_paths)}개 파일"
                self.pl_upload_zone.set_loading(label, pct)

    def _set_run_progress(self, pct):
        pct = max(0, min(100, int(pct)))
        self.pb_run["value"] = pct
        self.lbl_run_percent.configure(text=f"{pct}%")

    def _log(self, msg, color=None):
        self.log_queue.put((msg, color))
        self._log_history.append((msg, color))
        if len(self._log_history) > LOG_HISTORY_MAX:
            self._log_history = self._log_history[-LOG_HISTORY_MAX:]

    def _process_log_queue(self):
        try:
            while True:
                msg, color = self.log_queue.get_nowait()
                if self.log_container and self.log_win and self.log_win.winfo_exists():
                    self.log_container.add_log(msg, color)
                else:
                    self.unseen_log_count += 1
                    if hasattr(self, "btn_log_view"):
                        self.btn_log_view.text_val = f"로그 보기 ({self.unseen_log_count})"
                        self.btn_log_view._draw(self.btn_log_view.bg_color)
        except queue.Empty: pass
        self.after(50, self._process_log_queue)

    def _load_db_on_start(self):
        def task():
            try:
                self.db_records = load_database()
                self.after(0, lambda: self.db_card.set_loaded(self.db_records))
                self.after(0, lambda: self.btn_add_part.configure(state="normal"))
                br = category_breakdown_text(self.db_records)
                self._log(f"[DB] 총 {len(self.db_records)}건 로드 — 부품별: {br}", "green")
                self.after(0, lambda: self.lbl_pl_status.configure(
                    text="Part List: 업로드 대기 (오른쪽 박스에 PL 파일을 넣어주세요)",
                    fg=TEXT_SEC,
                ))
                self.after(0, lambda: self._set_pl_progress(0))
            except Exception as e:
                self.after(0, lambda: self.db_card.set_error(e))
                self._log(f"[오류] DB 연결 실패: {e}", "red")
        threading.Thread(target=task, daemon=True).start()

    def _load_part_list_worker(self):
        if not HAS_PART_LIST:
            self.pl_loading = False
            self.after(0, lambda: self.lbl_pl_status.configure(
                text="Part List: part_list_loader 없음", fg=TEXT_SEC))
            self.after(0, lambda: self._set_pl_progress(0))
            self.after(0, self._sync_vendor_button_state)
            return
        if not self.pl_paths:
            self.pl_lookup = {}
            self.pl_vendor_confirmed = False
            self.pl_loading = False
            self.after(0, lambda: self.lbl_pl_status.configure(
                text="Part List: 업로드된 파일 없음", fg=TEXT_SEC))
            self.after(0, lambda: self._set_pl_progress(0))
            self.after(0, self._sync_vendor_button_state)
            return
        try:
            def _pl_prog(done, total, _name):
                pct = int((done / total) * 100) if total else 100
                self.after(0, lambda p=pct: self._set_pl_progress(p))
            self.pl_lookup = load_part_list_from_paths(self.pl_paths, log=self._log, progress_cb=_pl_prog)
            src = f"선택 파일 {len(self.pl_paths)}개"
            self.pl_vendor_rank_by_loc = {
                k: v for k, v in self.pl_vendor_rank_by_loc.items() if k in self.pl_lookup
            }
            # 메모리 캐시 저장(동일 파일 조합이면 다음부터 즉시)
            sig = self._pl_signature(self.pl_paths)
            if sig:
                self._pl_mem_cache[sig] = {
                    "pl_lookup": self.pl_lookup,
                    "pl_vendor_rank_by_loc": dict(self.pl_vendor_rank_by_loc),
                }
                self._pl_last_sig = sig
            nloc = len(self.pl_lookup)
            nent = sum(len(v) for v in self.pl_lookup.values())
            self.after(0, lambda: self.lbl_pl_status.configure(
                text=f"Part List: 로드 완료(100%) - 로케이션 {nloc}개 / 후보 {nent}건 ({src})",
                fg=PRIMARY))
            self.after(0, lambda: self._set_pl_progress(100))
            done_label = os.path.basename(self.pl_paths[0]) if len(self.pl_paths) == 1 else f"{len(self.pl_paths)}개 파일"
            self.after(0, lambda l=done_label: self.pl_upload_zone.set_selected(l))
        except Exception as e:
            self.pl_lookup = {}
            self.pl_vendor_confirmed = False
            self.after(0, lambda: self.lbl_pl_status.configure(
                text=f"Part List: 오류 — {e}", fg=ERROR_COL))
            self._log(f"[PL] 로드 실패: {e}", "red")
            self.after(0, lambda: self._set_pl_progress(0))
            if self.pl_paths:
                err_label = os.path.basename(self.pl_paths[0]) if len(self.pl_paths) == 1 else f"{len(self.pl_paths)}개 파일"
                self.after(0, lambda l=err_label: self.pl_upload_zone.set_loading(l, 0))
        finally:
            self.pl_loading = False
            self.after(0, self._sync_vendor_button_state)

    def _stop_vendor_pulse(self):
        if getattr(self, "_vendor_pulse_job", None):
            try:
                self.after_cancel(self._vendor_pulse_job)
            except Exception:
                pass
            self._vendor_pulse_job = None
        if hasattr(self, "btn_vendor"):
            self.btn_vendor.set_attention_ring(False)

    def _vendor_pulse_tick(self):
        if (
            self.pl_vendor_confirmed
            or not self.var_pl_lookup.get()
            or not self.pl_lookup
            or self.pl_loading
        ):
            self._stop_vendor_pulse()
            return
        lit = not getattr(self, "_vendor_pulse_lit", False)
        self._vendor_pulse_lit = lit
        self.btn_vendor.set_attention_ring(lit)
        self._vendor_pulse_job = self.after(700, self._vendor_pulse_tick)

    def _start_vendor_pulse(self):
        if (
            self.pl_vendor_confirmed
            or not self.var_pl_lookup.get()
            or not self.pl_lookup
            or self.pl_loading
        ):
            return
        self._stop_vendor_pulse()
        self._vendor_pulse_lit = False
        self._vendor_pulse_tick()

    def _reset_vendor_button_to_pending(self):
        self.btn_vendor.bg_color = PRIMARY
        self.btn_vendor.hover_color = PRIMARY_HOV
        self.btn_vendor.text_color = "#ffffff"
        self.btn_vendor.icon = "▶"
        self.btn_vendor.text_val = "벤더 선택 — 로케이션별 1·2·3차 (Part List 파싱 후 필수)"
        self.btn_vendor.set_attention_ring(False)

    def _apply_vendor_button_done_style(self):
        self.btn_vendor.bg_color = SUCCESS
        self.btn_vendor.hover_color = "#047857"
        self.btn_vendor.text_color = "#ffffff"
        self.btn_vendor.icon = "✓"
        self.btn_vendor.text_val = "벤더 선택 완료 (다시 클릭해 수정 가능)"
        self.btn_vendor.set_attention_ring(False)

    def _sync_vendor_button_state(self):
        self._stop_vendor_pulse()
        if not hasattr(self, "btn_vendor"):
            return
        if self.pl_vendor_confirmed and self.pl_lookup and self.var_pl_lookup.get():
            self._apply_vendor_button_done_style()
        else:
            self._reset_vendor_button_to_pending()
            self._start_vendor_pulse()

    def _show_add_part_dialog(self):
        if not os.path.exists(DB_FILE):
            messagebox.showerror("DB 오류", f"Database 파일을 찾을 수 없습니다:\n{DB_FILE}")
            return
        AddPartDialog(self, on_success=self._reload_db_after_add)

    def _reload_db_after_add(self):
        def task():
            try:
                self.db_records = load_database()
                self.after(0, lambda: self.db_card.set_loaded(self.db_records))
                br = category_breakdown_text(self.db_records)
                self._log(f"[DB] 재로드 완료 — 총 {len(self.db_records)}건 · {br}", "green")
            except Exception as e:
                self.after(0, lambda: self.db_card.set_error(e))
                self._log(f"[오류] DB 재로드 실패: {e}", "red")
        threading.Thread(target=task, daemon=True).start()

    def _run(self):
        if not self.excel_path:
            messagebox.showwarning("파일 없음", "먼저 엑셀 파일을 업로드하세요.")
            return
        if not self.db_records:
            messagebox.showerror("DB 오류", "Database가 준비되지 않았습니다.")
            return
        if self.var_pl_lookup.get() and not self.pl_paths:
            messagebox.showinfo("안내", "Part List 보강을 사용하려면 오른쪽 박스에 PL 파일을 먼저 업로드하세요.")
            return
        if self.var_pl_lookup.get() and self.pl_loading:
            messagebox.showinfo("안내", "Part List를 읽는 중입니다. 상태가 '로케이션 ...'으로 바뀐 뒤 다시 실행해주세요.")
            return
        if (
            self.var_pl_lookup.get()
            and self.pl_paths
            and self.pl_lookup
            and not self.pl_loading
            and self.form_locations
            and not self.pl_vendor_confirmed
        ):
            messagebox.showwarning(
                "벤더 선택 필요",
                "Part List를 반영하려면 파란색 '벤더 선택' 버튼을 눌러 창을 연 뒤 내용을 확인하고 [완료]를 눌러 주세요.",
            )
            return

        files = list(self.batch_paths) if self.batch_paths else [self.excel_path]
        if self.var_validate.get():
            bad = []
            for fp in files:
                ok, msg = validate_workbook_quick(fp)
                if not ok:
                    bad.append(f"{os.path.basename(fp)}: {msg}")
            if bad:
                messagebox.showerror("검증 실패", "\n".join(bad))
                return

        opts = {
            "normalize_cell_text": self.var_norm.get(),
            "dry_run": self.var_dry_run.get(),
            "use_part_list_lookup": self.var_pl_lookup.get(),
            "pl_lookup": self.pl_lookup if self.var_pl_lookup.get() else {},
            "pl_vendor_rank_by_loc": self.pl_vendor_rank_by_loc if self.var_pl_lookup.get() else {},
        }
        self.batch_paths = []

        self.btn_run.configure(state="disabled")
        self.run_loading = True
        self._set_run_progress(0)
        self.status_bar.start_progress()

        def task():
            try:
                total_matched = total_checked = 0
                merged_unmatched = []
                n_files = max(1, len(files))
                for idx, fp in enumerate(files):
                    self._log(f"[System] 프로세스 시작: {fp}")
                    def _row_prog(cur, total, _msg):
                        ratio = (cur / total) if total else 1.0
                        overall = ((idx + ratio) / n_files) * 100.0
                        self.after(0, lambda p=overall: self._set_run_progress(p))
                    per_file_opts = dict(opts)
                    per_file_opts["progress_cb"] = _row_prog
                    matched, checked, result_path, unmatched = process_excel(
                        fp, self.db_records, self._log, per_file_opts
                    )
                    total_matched += matched
                    total_checked += checked
                    merged_unmatched.extend(unmatched)
                self.last_unmatched = merged_unmatched
                unmatched_count = len(merged_unmatched)
                self.after(0, lambda: self._set_run_progress(100))
                if n_files > 1:
                    self._log(
                        f"[안내] 일괄 처리: 미일치 {unmatched_count}건은 {n_files}개 파일을 합친 목록입니다.",
                        "yellow",
                    )
                
                def update_ui_success():
                    grouped = self._group_unmatched_by_part(merged_unmatched)
                    n_kinds = len(grouped)
                    self.status_bar.stop_progress(
                        f"완료 — 파일 {n_files}개 / 확인 {total_checked} / 일치 {total_matched} "
                        f"/ 미일치 {unmatched_count}곳 · 신규 후보 {n_kinds}종",
                        True,
                    )
                    if unmatched_count > 0:
                        self.btn_unmatched.configure(state="normal")
                        prompt = self._format_new_parts_prompt(grouped, max_locs_per_part=12)
                        if messagebox.askyesno("신규 부품 확인", prompt):
                            self._show_add_part_dialog()
                    else:
                        self.btn_unmatched.configure(state="disabled")
                    if self.var_open_result.get() and not self.var_dry_run.get():
                        self._open_result_folder()
                self.after(0, update_ui_success)
            except Exception as e:
                self._log(f"[오류] 실패: {e}", "red")
                self.after(0, lambda: self.status_bar.stop_progress(f"런타임 오류 발생: {e}", False))
                self.after(0, lambda: messagebox.showerror("처리 오류", str(e)))
            finally:
                self.run_loading = False
                self.after(0, lambda: self.btn_run.configure(state="normal"))

        threading.Thread(target=task, daemon=True).start()

    def _open_result_folder(self):
        result_dir = os.path.join(BASE_DIR, "Result")
        if os.path.exists(result_dir):
            try: os.startfile(result_dir)
            except: pass
        else:
            messagebox.showinfo("안내", "아직 Result 폴더가 없습니다.")

    def _group_unmatched_by_part(self, items):
        """
        last_unmatched = process_excel의 unmatched 목록(항목별: (fn, row, pn) 또는 (row, pn))
        을 '실제 부품·스펙' 기준으로 묶는다 (같은 부품이 여러 로케이션/행이면 1건).
        """
        groups = {}  # gkey -> {"display": str, "locs": set()}
        for it in items or []:
            if len(it) == 3:
                _, _, pn = it
                pn_str = str(pn).strip() if pn is not None else ""
                if not pn_str:
                    continue
            else:
                _, pn = it
                pn_str = str(pn).strip() if pn is not None else ""
                if not pn_str:
                    continue

            loc = _extract_location_label_from_unmatched(pn_str)
            if not loc:
                continue

            gkey, disp = _unmatched_group_key_and_display(pn_str)
            if not gkey:
                continue
            if gkey not in groups:
                groups[gkey] = {"display": disp or pn_str, "locs": set()}
            groups[gkey]["locs"].add(loc)
            if disp and len(disp) > len(groups[gkey].get("display") or ""):
                groups[gkey]["display"] = disp

        grouped = []
        for g in groups.values():
            locs_sorted = sorted(g["locs"], key=lambda s: (_norm_text(s), s))
            grouped.append((g["display"], locs_sorted))
        grouped.sort(key=lambda x: x[0])
        return grouped

    def _format_new_parts_prompt(self, grouped, max_locs_per_part=8):
        """
        grouped: [(pn, [locs...]), ...]
        """
        if not grouped:
            return "DB 미일치 항목이 있습니다.\n신규 부품 등록 창을 여시겠습니까?"

        total_spots = sum(len(locs) for _, locs in grouped)
        lines = []
        for pn, locs in grouped:
            if len(locs) <= max_locs_per_part:
                loc_txt = ", ".join(locs)
            else:
                loc_txt = ", ".join(locs[:max_locs_per_part]) + f" 외 {len(locs) - max_locs_per_part}개"
            lines.append(f"• {pn}\n  └ 로케이션: {loc_txt}")

        return (
            f"DB 미일치 신규 후보 부품 {len(grouped)}종 (총 {total_spots}개 로케이션).\n\n"
            + "\n".join(lines)
            + "\n\n신규 부품 등록 창을 여시겠습니까?"
        )

    def _show_unmatched(self):
        if not self.last_unmatched: return
        win = tk.Toplevel(self)
        win.title("미일치 항목")
        win.geometry("640x480")
        win.configure(bg=BG_MAIN)
        tk.Label(
            win,
            text=f"⚠ 미일치 항목 ({len(self.last_unmatched)}개)",
            bg=BG_MAIN,
            fg=WARN,
            font=("Segoe UI", 13, "bold"),
        ).pack(pady=16, padx=16, anchor="w")
        tb = scrolledtext.ScrolledText(win, bg=BG_CARD, fg=TEXT_PRI, font=("Consolas", 10), relief="flat")
        tb.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        for i, item in enumerate(self.last_unmatched, 1):
            if len(item) == 3:
                fn, r_num, pn = item
                tb.insert("end", f"{i:3d}. [{fn}] 행{r_num:4d}: {pn}\n")
            else:
                r_num, pn = item
                tb.insert("end", f"{i:3d}. 행{r_num:4d}: {pn}\n")
        tb.configure(state="disabled")

def main():
    try:
        from tkinterdnd2 import TkinterDnD, DND_FILES
        class AppDnD(TkinterDnD.Tk):
            def __init__(self):
                super().__init__()
                self.title("Auto Spec Program")
                self.state('zoomed')
                self.minsize(1024, 768)
                self.configure(bg=BG_MAIN)
                self.app_config = load_app_config()
                self.log_queue = queue.Queue()
                self._log_history = []
                self.db_records = []
                self.pl_lookup = {}
                self.pl_paths = []
                self.pl_vendor_rank_by_loc = {}
                self.form_locations = set()
                self.pl_loading = False
                self.run_loading = False
                self.log_win = None
                self.log_container = None
                self.unseen_log_count = 0
                self.excel_path = None
                self.last_unmatched = []
                self.batch_paths = []
                self.pl_vendor_confirmed = False
                self._vendor_pulse_job = None
                # Part List: 같은 실행 중 동일 파일 조합 재파싱 방지(결과 동일, 속도만 개선)
                self._pl_mem_cache = {}
                self._pl_last_sig = None

                App._build_ui(self)
                App._sync_vendor_button_state(self)
                App._bind_shortcuts(self)
                
                self.upload_zone.drop_target_register(DND_FILES)
                self.upload_zone.dnd_bind("<<Drop>>", App._on_drop_main.__get__(self))
                self.pl_upload_zone.drop_target_register(DND_FILES)
                self.pl_upload_zone.dnd_bind("<<Drop>>", App._on_drop_pl.__get__(self))
                
                App._process_log_queue(self)
                App._load_db_on_start(self)
                App._refresh_recent_combo(self)

            _select_file = App._select_file
            _set_file = App._set_file
            _extract_form_locations = App._extract_form_locations
            _log = App._log
            _run = App._run
            _open_result_folder = App._open_result_folder
            _group_unmatched_by_part = App._group_unmatched_by_part
            _format_new_parts_prompt = App._format_new_parts_prompt
            _show_unmatched = App._show_unmatched
            _show_add_part_dialog = App._show_add_part_dialog
            _reload_db_after_add = App._reload_db_after_add
            _process_log_queue = App._process_log_queue
            _load_db_on_start = App._load_db_on_start
            _bind_shortcuts = App._bind_shortcuts
            _persist_config = App._persist_config
            _refresh_recent_combo = App._refresh_recent_combo
            _on_recent_selected = App._on_recent_selected
            _select_batch_files = App._select_batch_files
            _paste_path_from_clipboard = App._paste_path_from_clipboard
            _export_log = App._export_log
            _clear_log = App._clear_log
            _copy_log_to_clipboard = App._copy_log_to_clipboard
            _open_log_window = App._open_log_window
            _parse_drop_paths = App._parse_drop_paths
            _on_drop_main = App._on_drop_main
            _on_drop_pl = App._on_drop_pl
            _select_part_list_files = App._select_part_list_files
            _pl_signature = App._pl_signature
            _set_part_list_files = App._set_part_list_files
            _open_vendor_selector = App._open_vendor_selector
            _set_pl_progress = App._set_pl_progress
            _set_run_progress = App._set_run_progress
            _load_part_list_worker = App._load_part_list_worker
            _sync_vendor_button_state = App._sync_vendor_button_state
            _stop_vendor_pulse = App._stop_vendor_pulse
            _vendor_pulse_tick = App._vendor_pulse_tick
            _start_vendor_pulse = App._start_vendor_pulse
            _reset_vendor_button_to_pending = App._reset_vendor_button_to_pending
            _apply_vendor_button_done_style = App._apply_vendor_button_done_style

        app = AppDnD()
    except ImportError:
        app = App()
    app.mainloop()

if __name__ == "__main__":
    main()
