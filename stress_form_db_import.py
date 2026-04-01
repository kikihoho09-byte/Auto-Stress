"""
Stress Analysis 양식(.xlsx)들을 스캔해 DB/Database.xlsx에 부품 행을 일괄 추가합니다.

  python stress_form_db_import.py --dry-run
  python stress_form_db_import.py --folder "db data"
  python stress_form_db_import.py --fix-cap-names              # CAP C열 짧은 이름 → 양식 기준 풀 문자열
  python stress_form_db_import.py --fix-cap-names --fix-cap-dry-run  # 저장 없이 건수만
  python stress_form_db_import.py --strip-cap-ripple           # CAP E열 '110mA [100Khz]' → '110mA'

양식 구조: 행 14~, G~L(Function), M~R(부품/로케이션), V열(SPEC.) — main.py 의 process_excel 과 동일 계열.
일부 파일(예: FVD, FDY 등 다른 시트/열 배치)은 0건으로 나올 수 있음 → 해당 파일은 열 오프셋 맞춤 필요.
"""
from __future__ import annotations

import argparse
import glob
import os
import re
import sys

ROOT = os.path.dirname(os.path.abspath(__file__))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from openpyxl import load_workbook

import main as app

try:
    from part_list_loader import normalize_el_cap_line
except ImportError:
    normalize_el_cap_line = None  # type: ignore

DB_CATEGORY_SCHEMA = app.DB_CATEGORY_SCHEMA
normalize_cell_text = app.normalize_cell_text
normalize_part_key = app.normalize_part_key
batch_append_parts_to_database = app.batch_append_parts_to_database
_is_mr_part_sheet_header_row = app._is_mr_part_sheet_header_row
find_db_sheet_name = app.find_db_sheet_name
DB_FILE = app.DB_FILE


def form_function_to_db_category(func_raw: str) -> str | None:
    """양식 Function 문자열 → AddPartDialog 카테고리 키(IC/MOSFET/DIODE/CAP/TR)."""
    if not func_raw:
        return None
    u = " ".join(str(func_raw).upper().replace("\r", "\n").replace("\n", " ").split())
    if "MOSFET" in u:
        return "MOSFET"
    if "RECTIFIER" in u or "DIODE" in u:
        return "DIODE"
    if "CAP" in u:
        return "CAP"
    if "TRANSISTOR" in u or re.search(r"\bTR\b", u):
        return "TR"
    if "FET" in u:
        return "MOSFET"
    if "IC" in u or "I.C" in u:
        return "IC"
    return None


def _looks_like_spec_token(t: str) -> bool:
    u = t.upper().replace(".", "")
    if re.match(r"^\d+V$", u) or re.match(r"^\d+(\.\d+)?UF$", u) or re.match(r"^\d+\*\d+", u):
        return True
    if re.match(r"^\d+[Vv](?:↓|↑)?$", t):
        return True
    if u.endswith("UF") and re.match(r"^\d", u):
        return True
    return False


def extract_pn_from_m_cell(m_val) -> str | None:
    """M열(첫 줄)에서 DB 부품명으로 쓸 MPN 후보 추출."""
    if m_val is None:
        return None
    first = str(m_val).split("\n")[0].strip()
    if not first:
        return None
    toks = first.split()
    candidates = []
    for tok in toks:
        t = tok.strip(" ,;:|")
        if len(t) < 3 or _looks_like_spec_token(t):
            continue
        if not re.search(r"\d", t) or not re.match(r"^[A-Z0-9.\-]+$", t, re.I):
            continue
        candidates.append(t)
    if candidates:
        return max(candidates, key=len).upper().replace(".", "")
    for tok in toks:
        t = tok.strip(" ,;:|")
        if _looks_like_spec_token(t) or len(t) < 2:
            continue
        if re.match(r"^[A-Z]", t, re.I):
            return t.upper().replace(".", "")
    return None


def cap_db_part_name_from_m(m13) -> str:
    """
    기존 Database CAP 시트 C열 포맷: 'KU 500V 39uF 10*50' (시리즈 + 전압 + uF + 사이즈).
    M열에서 RefDes 토큰만 제거한 뒤 part_list_loader 와 동일 규칙으로 조합한다.
    """
    if not m13 or normalize_el_cap_line is None:
        return ""
    line1 = str(m13).split("\n")[0].strip()
    if not line1:
        return ""
    toks = line1.split()
    while toks and app._is_schematic_ref_token(toks[-1]):
        toks.pop()
    cleaned = " ".join(toks).strip()
    if not cleaned:
        return ""
    return str(normalize_el_cap_line("", cleaned, "") or "").strip()


def _tail_after_pn_mpn(m_cell, pn: str) -> str:
    if not m_cell or not pn:
        return ""
    m0 = str(m_cell).split("\n")[0].strip()
    if not m0:
        return ""
    up, pnu = m0.upper(), pn.upper()
    idx = up.find(pnu)
    if idx >= 0:
        tail = (m0[:idx] + m0[idx + len(pn) :]).strip()
    else:
        tail = m0
        for tok in m0.split():
            if tok.strip(" ,;:|").upper().replace(".", "") == pnu.replace(".", ""):
                tail = m0.replace(tok, " ", 1).strip()
                break
    return " ".join(tail.split())


def _clean_spec_value(v) -> str:
    if v is None:
        return ""
    s = " ".join(str(v).replace("\n", " ").replace("\r", " ").split()).strip()
    s = s.replace("↓", "").replace("↑", "").strip()
    return s


def _strip_cap_ripple_brackets(s) -> str:
    """CAP 리플 전류 셀: '110mA [100Khz]' → '110mA' (주파수 괄호 제거)."""
    s = _clean_spec_value(s)
    if not s:
        return ""
    s = re.sub(r"\s*\[[^\]]+\]", "", s)
    return " ".join(s.split()).strip()


def _norm_factor_label(v) -> str:
    s = _clean_spec_value(v).upper()
    s = re.sub(r"[\s\-\._/()]+", "", s)
    return s


def _build_specs_from_block(category: str, block_rows: list[list]) -> list[str]:
    _, n = DB_CATEGORY_SCHEMA[category]
    specs = [""] * n
    for vals in block_rows:
        if not vals or len(vals) < 16:
            continue
        factor = _norm_factor_label(vals[12])  # S열(Test Factor)
        val = _clean_spec_value(vals[15])       # V열(SPEC)
        if not val:
            continue
        if category == "MOSFET":
            if "VDS" in factor:
                specs[0] = specs[0] or val
            elif "VGS" in factor:
                specs[1] = specs[1] or val
            elif "IDS" in factor or factor == "ID":
                specs[2] = specs[2] or val
        elif category == "DIODE":
            if "VRR" in factor:
                specs[0] = specs[0] or val
            elif "IFSM" in factor:
                specs[2] = specs[2] or val
            elif "IF" in factor:
                specs[1] = specs[1] or val
        elif category == "TR":
            if "VCB" in factor:
                specs[0] = specs[0] or val
            elif "VEB" in factor:
                specs[1] = specs[1] or val
            elif factor == "ID" or "ID" in factor:
                specs[2] = specs[2] or val
            elif "ICP" in factor:
                specs[3] = specs[3] or val
        elif category == "CAP":
            if "RIPPLE" in factor or "IRIPPLE" in factor or factor.startswith("IR"):
                specs[1] = specs[1] or val
            elif ("VPP" in factor or "V" in factor) and not specs[0]:
                specs[0] = val
        elif category == "IC":
            if "VCC" in factor or not specs[0]:
                specs[0] = specs[0] or val
    return specs


def _finalize_cap_specs(specs: list[str], n: int) -> None:
    if n == 2 and len(specs) > 1 and specs[1]:
        specs[1] = _strip_cap_ripple_brackets(specs[1])


def build_specs_for_category(category: str, m_cell, v_cell, block_rows: list[list]) -> list[str]:
    _, n = DB_CATEGORY_SCHEMA[category]
    specs = _build_specs_from_block(category, block_rows)
    v = _clean_spec_value(v_cell)
    pn = extract_pn_from_m_cell(m_cell) or ""
    tail = _tail_after_pn_mpn(m_cell, pn)

    if n == 1:
        specs[0] = specs[0] or v or tail[:200]
        return specs

    if not specs[0]:
        specs[0] = v

    if not tail:
        if category == "CAP":
            _finalize_cap_specs(specs, n)
        return specs

    words = tail.split()
    if n == 2:
        # CAP의 Iripple은 mA 등 전류값만 채운다(용량/사이즈 문자열 오입력 방지)
        if not specs[1]:
            m_curr = re.search(r"\b\d+(?:\.\d+)?\s*mA\b", tail, re.I)
            if m_curr:
                specs[1] = m_curr.group(0).replace(" ", "")
        if not specs[0] and not v:
            specs[0] = words[0] if words else tail[:120]
        if category == "CAP":
            _finalize_cap_specs(specs, n)
        return specs

    need = n - 1
    if len(words) >= need:
        for i in range(need):
            if not specs[1 + i]:
                specs[1 + i] = words[i][:120]
    else:
        if not specs[1]:
            specs[1] = tail[:160]
        if need >= 2:
            if not specs[2]:
                specs[2] = tail[160:320] if n >= 4 else tail[160:300]
        if need >= 3:
            if not specs[3]:
                specs[3] = tail[320:480] if n == 4 else ""
    if category == "CAP":
        _finalize_cap_specs(specs, n)
    return specs


def collect_rows_from_workbook(path: str) -> list[dict]:
    """read_only 워크북은 iter_rows 로만 읽는다(cell random access는 매우 느림)."""
    src = os.path.basename(path)
    out: list[dict] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        mr = ws.max_row or 14
        max_r = min(mr, 14 + 1200)
        row_map: dict[int, list] = {}
        for r_idx, row_cells in enumerate(
            ws.iter_rows(
                min_row=14,
                max_row=max_r,
                min_col=7,
                max_col=22,
                values_only=True,
            ),
            start=14,
        ):
            vals = list(row_cells)
            if len(vals) < 16:
                vals = list(vals) + [None] * (16 - len(vals))
            row_map[r_idx] = vals
        # G=7 … V=22 스냅샷을 기반으로 블록 단위 해석
        for r_idx in sorted(row_map.keys()):
            vals = row_map[r_idx]
            func = None
            for j in range(0, 6):
                v = vals[j]
                if v is not None and str(v).strip():
                    func = str(v).strip()
                    break
            if not func:
                continue
            cat = form_function_to_db_category(func)
            if not cat:
                continue
            m_slice = vals[6:12]
            m_values = []
            for v in m_slice:
                if v:
                    m_values.append(normalize_cell_text(v))
            if not m_values:
                continue
            joined = " / ".join(m_values)
            if _is_mr_part_sheet_header_row(m_values, joined):
                continue
            m13 = m_slice[0] if m_slice else None
            short_pn = extract_pn_from_m_cell(m13)
            v22 = vals[15]
            block_rows = [row_map.get(rr, [None] * 16) for rr in range(r_idx, min(r_idx + 7, max_r + 1))]
            specs = build_specs_for_category(cat, m13, v22, block_rows)
            if cat == "MOSFET" and (not specs[1] or not specs[2]):
                continue
            if cat == "CAP" and not specs[1]:
                continue
            if cat == "CAP":
                part_name = cap_db_part_name_from_m(m13)
                if not part_name:
                    continue
            else:
                if not short_pn:
                    continue
                part_name = short_pn
            out.append(
                {
                    "category": cat,
                    "part_name": part_name,
                    "specs": specs,
                    "_src": src,
                    "_row": r_idx,
                    "_m13": m13,
                    "_short_pn": short_pn if cat == "CAP" else None,
                }
            )
    finally:
        wb.close()
    return out


def _norm_cap_spec_key(v) -> str:
    t = _clean_spec_value(v).lower()
    return re.sub(r"\s+", "", t)


def _norm_cap_ripple_key(v) -> str:
    """조회/중복 비교용: 리플 문자열에서 괄호 블록 제거 후 정규화."""
    return _norm_cap_spec_key(_strip_cap_ripple_brackets(v))


def _cap_series_key_from_c_cell(cval) -> str:
    """CAP 시트 C열에서 짧은(시리즈만) 행과 매칭할 시리즈 키."""
    s = str(cval or "").strip()
    if not s:
        return ""
    if "UF" in s.upper() or "μF" in s or "ΜF" in s.upper():
        return normalize_part_key(s)
    toks = s.split()
    return normalize_part_key(toks[0]) if toks else ""


def _cap_c_cell_needs_pl_style(cval) -> bool:
    """기존 DB처럼 '시리즈 전압 uF 사이즈' 풀 문자열이 아닌 짧은 C열만 교정."""
    s = str(cval or "").strip()
    if not s:
        return False
    u = s.upper()
    if "UF" in u or "μF" in s:
        return False
    if re.search(r"\d+\s*[*xX]\s*\d+", s):
        return False
    return True


def _cap_lookup_key_row(short_pn, d_spec, e_spec) -> tuple[str, str, str]:
    sk = normalize_part_key(str(short_pn or "").strip())
    return (sk, _norm_cap_spec_key(d_spec), _norm_cap_ripple_key(e_spec))


def build_cap_part_name_lookup_from_files(files: list[str]) -> tuple[dict[tuple[str, str, str], str], list[tuple]]:
    cap_map: dict[tuple[str, str, str], str] = {}
    conflicts: list[tuple] = []
    for fp in files:
        try:
            rows = collect_rows_from_workbook(fp)
        except Exception:
            continue
        for e in rows:
            if e.get("category") != "CAP":
                continue
            full = str(e.get("part_name") or "").strip()
            if not full:
                continue
            specs = e.get("specs") or []
            if len(specs) < 2:
                continue
            short = e.get("_short_pn") or extract_pn_from_m_cell(e.get("_m13"))
            if not short:
                p0 = full.split()
                short = p0[0] if p0 else ""
            key = _cap_lookup_key_row(short, specs[0], specs[1])
            if not key[0]:
                continue
            prev = cap_map.get(key)
            if prev and prev != full:
                conflicts.append((key, prev, full))
            else:
                cap_map[key] = full
    return cap_map, conflicts


def fix_cap_part_names_in_database(
    folder: str,
    limit_files: int = 0,
    dry_run: bool = False,
    single_file: str = "",
) -> dict:
    """CAP 시트에서 C열이 시리즈만 있는 행을 양식 기준 풀 part 문자열로 덮어쓴다."""
    if single_file and os.path.isfile(single_file):
        files = [single_file]
    else:
        pattern = os.path.join(folder, "*.xlsx")
        files = sorted(glob.glob(pattern), key=os.path.getsize)
        if limit_files > 0:
            files = files[:limit_files]
    cap_map, conflicts = build_cap_part_name_lookup_from_files(files)
    result = {
        "files_scanned": len(files),
        "lookup_size": len(cap_map),
        "conflicts": len(conflicts),
        "updated": 0,
        "skipped_ok": 0,
        "missing_key": 0,
    }
    if conflicts:
        print(f"[fix-cap] 조회 충돌 {len(conflicts)}건 (같은 시리즈+D+E에 서로 다른 풀이름) — 마지막 것만 유지", flush=True)

    wb = load_workbook(DB_FILE, read_only=False, keep_vba=False)
    try:
        sn = find_db_sheet_name(wb, "CAP")
        if not sn:
            raise RuntimeError("CAP 시트를 찾을 수 없습니다.")
        ws = wb[sn]
        missing_samples: list[tuple] = []
        for r in range(3, (ws.max_row or 2) + 1):
            c = ws.cell(row=r, column=3).value
            d = ws.cell(row=r, column=4).value
            e = ws.cell(row=r, column=5).value
            if not c or not str(c).strip():
                continue
            if not _cap_c_cell_needs_pl_style(c):
                result["skipped_ok"] += 1
                continue
            key = (
                _cap_series_key_from_c_cell(c),
                _norm_cap_spec_key(d),
                _norm_cap_ripple_key(e),
            )
            full = cap_map.get(key)
            if not full:
                result["missing_key"] += 1
                if len(missing_samples) < 12:
                    missing_samples.append((r, c, d, e, key))
                continue
            if str(c).strip() != full:
                result["updated"] += 1
                if not dry_run:
                    ws.cell(row=r, column=3).value = full
        if not dry_run:
            wb.save(DB_FILE)
        result["missing_samples"] = missing_samples
    finally:
        wb.close()
    return result


def strip_cap_ripple_in_database(dry_run: bool = False) -> dict:
    """CAP 시트 E열 리플 전류에서 '[100Khz]' 등 대괄호 표기만 제거한다."""
    result = {"updated": 0, "checked": 0}
    wb = load_workbook(DB_FILE, read_only=False, keep_vba=False)
    try:
        sn = find_db_sheet_name(wb, "CAP")
        if not sn:
            raise RuntimeError("CAP 시트를 찾을 수 없습니다.")
        ws = wb[sn]
        for r in range(3, (ws.max_row or 2) + 1):
            e = ws.cell(row=r, column=5).value
            if e is None or not str(e).strip():
                continue
            result["checked"] += 1
            old = _clean_spec_value(e)
            new_e = _strip_cap_ripple_brackets(e)
            if new_e != old:
                result["updated"] += 1
                if not dry_run:
                    ws.cell(row=r, column=5).value = new_e
        if not dry_run and result["updated"]:
            wb.save(DB_FILE)
    finally:
        wb.close()
    return result


def dedupe_entries(entries: list[dict]) -> list[dict]:
    seen: set[tuple[str, str]] = set()
    uniq: list[dict] = []
    for e in entries:
        k = (e["category"], normalize_part_key(e["part_name"]))
        if k in seen:
            continue
        seen.add(k)
        ee = {k2: v2 for k2, v2 in e.items() if not k2.startswith("_")}
        uniq.append(ee)
    return uniq


def main():
    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(encoding="utf-8")
        except Exception:
            pass
    ap = argparse.ArgumentParser(description="Stress 양식 → Database.xlsx 일괄 등록")
    ap.add_argument(
        "--folder",
        default=os.path.join(ROOT, "db data"),
        help="양식 xlsx 가 있는 폴더",
    )
    ap.add_argument("--dry-run", action="store_true", help="DB에 쓰지 않고 집계만")
    ap.add_argument("--limit-files", type=int, default=0, help="테스트용: 최대 n개 파일만")
    ap.add_argument("--file", default="", help="단일 xlsx만 처리(폴더 대신)")
    ap.add_argument(
        "--yes",
        action="store_true",
        help="확인 프롬프트 없이 DB에 기록(자동 실행용)",
    )
    ap.add_argument(
        "--fix-cap-names",
        action="store_true",
        help="CAP 시트 C열(시리즈만 있는 행)을 db data 양식 기준 풀 문자열로만 교정",
    )
    ap.add_argument(
        "--fix-cap-dry-run",
        action="store_true",
        help="--fix-cap-names 와 함께: 저장 없이 교정 대상 행 수만 표시",
    )
    ap.add_argument(
        "--strip-cap-ripple",
        action="store_true",
        help="CAP 시트 E열에서 리플 전류 뒤 [주파수] 대괄호 제거",
    )
    ap.add_argument(
        "--strip-cap-ripple-dry-run",
        action="store_true",
        help="--strip-cap-ripple 와 함께: 저장 없이 건수만",
    )
    args = ap.parse_args()

    if not os.path.isfile(DB_FILE):
        print(f"Database 없음: {DB_FILE}")
        sys.exit(1)

    if args.strip_cap_ripple:
        print(strip_cap_ripple_in_database(dry_run=args.strip_cap_ripple_dry_run))
        return

    if args.fix_cap_names:
        res = fix_cap_part_names_in_database(
            args.folder,
            limit_files=args.limit_files,
            dry_run=args.fix_cap_dry_run,
            single_file=args.file,
        )
        print(res)
        if res.get("missing_samples"):
            print("[fix-cap] 맵에 없어 건너뜀(샘플):", flush=True)
            for t in res["missing_samples"]:
                print(f"  row {t[0]} C={t[1]!r} key={t[4]}", flush=True)
        return

    if args.file:
        if not os.path.isfile(args.file):
            print(f"파일 없음: {args.file}")
            sys.exit(1)
        files = [args.file]
    else:
        if not os.path.isdir(args.folder):
            print(f"폴더 없음: {args.folder}")
            sys.exit(1)
        pattern = os.path.join(args.folder, "*.xlsx")
        files = sorted(glob.glob(pattern), key=os.path.getsize)
    if args.limit_files > 0:
        files = files[: args.limit_files]

    all_raw: list[dict] = []
    print(f"처리 파일 {len(files)}개 (작은 파일부터)…", flush=True)
    for fp in files:
        try:
            print(f"  … {os.path.basename(fp)}", flush=True)
            rows = collect_rows_from_workbook(fp)
            all_raw.extend(rows)
            print(f"  읽음 {os.path.basename(fp)}: {len(rows)}행 (부품 후보)", flush=True)
        except Exception as ex:
            print(f"  건너뜀 {os.path.basename(fp)}: {ex}", flush=True)

    merged = dedupe_entries(all_raw)
    print(f"\n합계: 파일 {len(files)}개, 원시 {len(all_raw)}행 → 중복 제거 후 {len(merged)}건")

    if args.dry_run:
        from collections import Counter

        c = Counter(e["category"] for e in merged)
        print("카테고리별:", dict(c))
        for sample in merged[:15]:
            print(f"  [{sample['category']}] {sample['part_name']} | specs={sample['specs']}")
        if len(merged) > 15:
            print(f"  ... 외 {len(merged) - 15}건")
        print("\n--dry-run 이므로 Database.xlsx 는 변경하지 않았습니다.")
        print("실제 반영: python stress_form_db_import.py --yes")
        return

    if not args.yes:
        try:
            ans = input(f"\nDatabase.xlsx 에 최대 {len(merged)}건 추가를 시도합니다. 진행하려면 yes 입력: ")
        except EOFError:
            ans = ""
        if ans.strip().lower() != "yes":
            print("취소됨.")
            return

    res = batch_append_parts_to_database(merged, log_func=print)
    print(res)


if __name__ == "__main__":
    main()
